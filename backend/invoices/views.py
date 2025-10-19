"""
Views para el módulo de Invoices.
Maneja CRUD de facturas, upload de archivos, matching y estadísticas.
"""

from rest_framework import viewsets, status
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.db.models import Q, Sum, Count
from django.utils import timezone
from django.core.files.storage import storages
from django.http import FileResponse, HttpResponse
from decimal import Decimal
from datetime import datetime, date
import uuid
import os
import re
import zipfile
from io import BytesIO
import logging

# Helper function to get the configured storage
def get_storage():
    """Returns the configured default storage (FileSystem or Cloudinary)"""
    return storages['default']

from .models import Invoice, UploadedFile, Dispute, CreditNote, DisputeEvent
from ots.models import OT
from catalogs.models import Provider
from .serializers import (
    InvoiceListSerializer,
    InvoiceDetailSerializer,
    InvoiceCreateSerializer,
    InvoiceUpdateSerializer,
    InvoiceStatsSerializer,
    UploadedFileSerializer,
    DisputeListSerializer,
    DisputeDetailSerializer,
    DisputeCreateSerializer,
    DisputeUpdateSerializer,
    DisputeEventSerializer,
    CreditNoteListSerializer,
    CreditNoteDetailSerializer,
    CreditNoteCreateSerializer,
    CreditNoteUpdateSerializer,
)
from common.permissions import IsJefeOperaciones


class InvoiceViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestión de facturas.
    
    Endpoints:
    - GET /invoices/ - Listar facturas
    - POST /invoices/ - Crear factura manual
    - GET /invoices/{id}/ - Detalle de factura
    - PUT/PATCH /invoices/{id}/ - Actualizar factura
    - DELETE /invoices/{id}/ - Eliminar (soft delete) factura
    - POST /invoices/upload/ - Upload masivo de facturas
    - GET /invoices/pending/ - Facturas pendientes de revisión
    - GET /invoices/stats/ - Estadísticas
    - POST /invoices/{id}/assign_ot/ - Asignar OT manualmente
    """
    
    permission_classes = [IsAuthenticated]
    queryset = Invoice.objects.filter(is_deleted=False).select_related(
        'ot', 'proveedor', 'uploaded_file'
    )
    lookup_value_regex = r'[0-9]+'
    
    def get_serializer_class(self):
        if self.action == 'list':
            return InvoiceListSerializer
        elif self.action in ['create', 'upload']:
            return InvoiceCreateSerializer
        elif self.action in ['update', 'partial_update', 'assign_ot']:
            return InvoiceUpdateSerializer
        elif self.action == 'stats':
            return InvoiceStatsSerializer
        return InvoiceDetailSerializer
    
    def get_queryset(self):
        """
        Filtrado avanzado de facturas.
        Query params:
        - estado_provision: pendiente, provisionada, rechazada
        - estado_facturacion: pendiente, facturada
        - requiere_revision: true, false
        - ot_number: número de OT
        - proveedor_nombre: búsqueda parcial
        - fecha_desde: YYYY-MM-DD
        - fecha_hasta: YYYY-MM-DD
        - search: búsqueda general
        """
        queryset = super().get_queryset()
        
        # Filtros
        estado_provision = self.request.query_params.get('estado_provision')
        if estado_provision:
            queryset = queryset.filter(estado_provision=estado_provision)
        
        estado_facturacion = self.request.query_params.get('estado_facturacion')
        if estado_facturacion:
            queryset = queryset.filter(estado_facturacion=estado_facturacion)
        
        requiere_revision = self.request.query_params.get('requiere_revision')
        if requiere_revision:
            value = requiere_revision.lower() == 'true'
            queryset = queryset.filter(requiere_revision=value)
        
        ot_number = self.request.query_params.get('ot_number')
        if ot_number:
            queryset = queryset.filter(ot_number__icontains=ot_number)
        
        proveedor = self.request.query_params.get('proveedor_nombre')
        if proveedor:
            queryset = queryset.filter(proveedor_nombre__icontains=proveedor)
        
        ot_id = self.request.query_params.get('ot') or self.request.query_params.get('ot_id')
        if ot_id:
            queryset = queryset.filter(ot_id=ot_id)

        fecha_desde = self.request.query_params.get('fecha_desde')
        if fecha_desde:
            queryset = queryset.filter(fecha_emision__gte=fecha_desde)
        
        fecha_hasta = self.request.query_params.get('fecha_hasta')
        if fecha_hasta:
            queryset = queryset.filter(fecha_emision__lte=fecha_hasta)
        
        # Búsqueda general
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(numero_factura__icontains=search) |
                Q(proveedor_nombre__icontains=search) |
                Q(ot_number__icontains=search) |
                Q(notas__icontains=search)
            )
        
        return queryset.distinct()
    
    @action(detail=True, methods=['get'], url_path='file')
    def retrieve_file(self, request, pk=None):
        """Permite descargar o previsualizar el archivo original de la factura."""
        from django.shortcuts import redirect
        from django.conf import settings

        invoice = self.get_object()

        if not invoice.uploaded_file:
            return Response(
                {'detail': 'La factura no tiene archivo asociado.'},
                status=status.HTTP_404_NOT_FOUND
            )

        storage_path = invoice.uploaded_file.path

        # If using Cloudinary, redirect directly to Cloudinary URL (fast!)
        if getattr(settings, 'USE_CLOUDINARY', False):
            try:
                storage = get_storage()
                cloudinary_url = storage.url(storage_path)
                return redirect(cloudinary_url)
            except Exception as exc:
                return Response(
                    {'detail': f'Error al obtener URL del archivo: {exc}'},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )

        # For local filesystem, serve file normally
        if not get_storage().exists(storage_path):
            return Response(
                {'detail': 'Archivo no encontrado en el almacenamiento.'},
                status=status.HTTP_404_NOT_FOUND
            )

        try:
            file_handle = get_storage().open(storage_path, 'rb')
        except Exception as exc:  # pragma: no cover
            return Response(
                {'detail': f'No se pudo abrir el archivo: {exc}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        # Generar nombre de archivo amigable usando short_name del cliente si existe
        filename = self._generate_friendly_filename(invoice)
        if not filename:
            # Fallback al nombre original
            filename = invoice.uploaded_file.filename or storage_path.split('/')[-1]

        content_type = invoice.uploaded_file.content_type or 'application/octet-stream'

        response = FileResponse(file_handle, content_type=content_type)

        download_flag = str(request.query_params.get('download', '')).lower()
        disposition = 'attachment' if download_flag in ('1', 'true', 'yes') else 'inline'
        response['Content-Disposition'] = f'{disposition}; filename="{filename}"'
        response['Content-Length'] = invoice.uploaded_file.size
        response['Access-Control-Expose-Headers'] = 'Content-Disposition'

        return response
    
    def _generate_friendly_filename(self, invoice):
        """
        Genera un nombre de archivo amigable para la factura.
        Formato: FACTURA PROVEEDOR NUMERO ALIAS OT.pdf
        (con espacios, sin guiones)
        
        Ejemplo: FACTURA MAERSK INV2024001 SIMAN OT2024123.pdf
        """
        try:
            # Construir nombre de archivo limpio
            proveedor_name = invoice.proveedor.nombre if invoice.proveedor else invoice.proveedor_nombre
            proveedor_name = re.sub(r'[^\w\s]', '', proveedor_name).strip()[:30]

            numero_factura = re.sub(r'[^\w]', '', invoice.numero_factura)[:30]
            
            cliente_short = ''
            if invoice.ot and invoice.ot.cliente and invoice.ot.cliente.short_name:
                cliente_short = re.sub(r'[^\w\s]', '', invoice.ot.cliente.short_name).strip()[:20]
            
            ot_number = ''
            if invoice.ot:
                ot_number = re.sub(r'[^\w]', '', invoice.ot.numero_ot)[:20]

            # Obtener extensión
            original_filename = invoice.uploaded_file.filename or 'invoice.pdf'
            _, ext = os.path.splitext(original_filename)
            if not ext:
                ext = '.pdf'
            
            # Construir nombre con espacios (sin guiones)
            parts = ['FACTURA', proveedor_name, numero_factura]
            if cliente_short:
                parts.append(cliente_short)
            if ot_number:
                parts.append(ot_number)
            
            filename = ' '.join(filter(None, parts)) + ext

            # Si es muy largo (>150 chars), usar formato corto
            if len(filename) > 150:
                filename = f"{numero_factura} {ot_number}{ext}" if ot_number else f"{numero_factura}{ext}"
            
            return filename
            
        except Exception as e:
            # En caso de error, retornar None para usar el fallback
            logger = logging.getLogger(__name__)
            logger.warning(f"Error generando nombre amigable para factura {invoice.id}: {e}")
            return None

    @action(detail=False, methods=['post'], permission_classes=[IsAuthenticated])
    def upload(self, request):
        """
        Upload múltiple de facturas con reconocimiento automático usando patrones.
        
        Proceso:
        1. Guardar archivo y calcular hash
        2. Extraer texto del PDF
        3. Aplicar patrones del proveedor + genéricos
        4. Crear factura con campos auto-detectados
        5. Si se detecta contenedor, intentar match con OT
        """
        from .parsers.pdf_extractor import PDFExtractor
        from .parsers.pattern_service import PatternApplicationService
        
        logger = logging.getLogger(__name__)
        
        files = request.FILES.getlist('files[]')
        proveedor_id = request.data.get('proveedor_id')
        tipo_costo = request.data.get('tipo_costo', 'OTRO')
        # CAMBIO: auto_parse desactivado por defecto para evitar timeouts
        auto_parse = request.data.get('auto_parse', 'false').lower() == 'true'
        
        if not files:
            return Response(
                {'error': 'No se enviaron archivos'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if not proveedor_id:
            return Response(
                {'error': 'Debe seleccionar un proveedor'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Verificar que el proveedor existe
        try:
            proveedor = Provider.objects.get(id=proveedor_id)
        except Provider.DoesNotExist:
            return Response(
                {'error': 'Proveedor no encontrado'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Inicializar servicios
        pdf_extractor = PDFExtractor()
        pattern_service = PatternApplicationService(provider_id=int(proveedor_id))
        
        results = {
            'success': [],
            'errors': [],
            'duplicates': [],
            'patterns_used': pattern_service.get_patterns_summary(),  # Para mostrar en frontend
        }
        
        for file in files:
            try:
                # Leer archivo UNA SOLA VEZ al inicio
                file.seek(0)
                file_content = file.read()

                # Calcular hash
                import hashlib
                file_hash = hashlib.sha256(file_content).hexdigest()

                # Verificar duplicado: buscar archivo por hash
                existing_file = UploadedFile.objects.filter(sha256=file_hash).first()

                if existing_file:
                    # Verificar si hay facturas ACTIVAS asociadas a este archivo
                    # Si todas las facturas fueron eliminadas (is_deleted=True), permitir re-subir
                    active_invoices = Invoice.objects.filter(
                        uploaded_file=existing_file,
                        is_deleted=False
                    ).exists()

                    if active_invoices:
                        results['duplicates'].append({
                            'filename': file.name,
                            'reason': f'Archivo duplicado con factura activa (SHA256: {file_hash[:8]}...)'
                        })
                        continue
                    # Si no hay facturas activas, permitir re-subir (reutilizar el mismo archivo)

                # Guardar archivo (solo si no existe) o reutilizar el existente
                if not existing_file:
                    from django.core.files.base import ContentFile
                    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
                    safe_filename = f"{timestamp}_{file.name}"

                    # Crear ContentFile desde bytes ya leídos
                    file_obj = ContentFile(file_content, name=safe_filename)
                    path = get_storage().save(f'invoices/{safe_filename}', file_obj)

                    uploaded_file = UploadedFile.objects.create(
                        filename=file.name,
                        path=path,
                        sha256=file_hash,
                        size=len(file_content),
                        content_type=file.content_type
                    )
                else:
                    # Reutilizar archivo existente (fue eliminado anteriormente)
                    uploaded_file = existing_file

                # Datos extraídos (con valores por defecto)
                extracted_data = {
                    'numero_factura': None,
                    'fecha_emision': timezone.localdate(),
                    'monto': Decimal('0.00'),
                    'numero_contenedor': None,
                    'mbl': None,
                    'ot_matched': None,
                    'confidence': 0.0,
                    'extraction_details': {},
                }

                # Auto-parsing con patrones (usando contenido ya leído)
                if auto_parse and file.content_type == 'application/pdf':
                    try:
                        # 1. Extraer texto del PDF (ya tenemos el contenido)
                        pdf_result = pdf_extractor.extract(file_content)
                        text = pdf_extractor.text
                        
                        # 2. Aplicar patrones
                        if text:
                            pattern_results = pattern_service.apply_patterns(text)
                            
                            # 3. Mapear resultados a campos de factura
                            if 'numero_factura' in pattern_results:
                                extracted_data['numero_factura'] = pattern_results['numero_factura']['value']
                                extracted_data['confidence'] = pattern_results['numero_factura']['confidence']
                            
                            if 'monto_total' in pattern_results:
                                extracted_data['monto'] = pattern_results['monto_total']['value']
                            
                            if 'fecha_emision' in pattern_results:
                                fecha = pattern_results['fecha_emision']['value']
                                if fecha:
                                    if isinstance(fecha, datetime):
                                        extracted_data['fecha_emision'] = fecha.date()
                                    elif isinstance(fecha, date):
                                        extracted_data['fecha_emision'] = fecha
                                    else:
                                        # Intentar convertir string a fecha
                                        try:
                                            from dateutil import parser
                                            parsed_date = parser.parse(str(fecha))
                                            extracted_data['fecha_emision'] = parsed_date.date()
                                        except:
                                            logger.warning(f"No se pudo parsear fecha: {fecha}")
                                            # Mantener fecha por defecto
                            
                            if 'numero_contenedor' in pattern_results:
                                extracted_data['numero_contenedor'] = pattern_results['numero_contenedor']['value']
                            
                            if 'mbl' in pattern_results:
                                extracted_data['mbl'] = pattern_results['mbl']['value']
                            
                            # 4. Intentar matching automático con OT
                            # Prioridad 1: MBL (más específico)
                            # Prioridad 2: Contenedor (menos específico)
                            matched_ot = None
                            match_method = None
                            
                            if extracted_data['mbl']:
                                mbl = extracted_data['mbl'].strip()
                                matched_ot = OT.objects.filter(
                                    master_bl__icontains=mbl,
                                    is_deleted=False
                                ).first()
                                
                                if matched_ot:
                                    match_method = 'MBL'
                                    logger.info(f"✓ OT matched por MBL '{mbl}': {matched_ot.numero_ot}")
                            
                            # Si no encontró por MBL, intentar con contenedor
                            if not matched_ot and extracted_data['numero_contenedor']:
                                contenedor = extracted_data['numero_contenedor'].strip()
                                # Buscar en el array de contenedores (campo JSONField)
                                matched_ot = OT.objects.filter(
                                    contenedores__contains=[contenedor],
                                    is_deleted=False
                                ).first()
                                
                                # Si no encuentra exacto, buscar parcial en el array
                                if not matched_ot:
                                    # Búsqueda parcial eficiente usando icontains en el campo JSON.
                                    # Esto es mucho más rápido que iterar todas las OTs en Python.
                                    matched_ot = OT.objects.filter(
                                        contenedores__icontains=contenedor,
                                        is_deleted=False
                                    ).first()
                                
                                if matched_ot:
                                    match_method = 'Contenedor'
                                    logger.info(f"✓ OT matched por Contenedor '{contenedor}': {matched_ot.numero_ot}")
                            
                            # Guardar resultado del matching
                            if matched_ot:
                                extracted_data['ot_matched'] = matched_ot.numero_ot
                                extracted_data['match_method'] = match_method
                            else:
                                if extracted_data['mbl'] or extracted_data['numero_contenedor']:
                                    logger.warning(
                                        f"⚠ No se encontró OT para MBL='{extracted_data['mbl']}' "
                                        f"o Contenedor='{extracted_data['numero_contenedor']}'"
                                    )
                            
                            # Guardar todos los detalles
                            extracted_data['extraction_details'] = pattern_results
                    
                    except Exception as e:
                        logger.error(f"Error en auto-parsing para {file.name}: {e}")
                        # Continuar con valores por defecto
                
                # Crear Invoice
                # Si no se extrajo número de factura, usar temporal
                numero_factura = extracted_data['numero_factura']
                if not numero_factura:
                    numero_factura = f"TEMP-{uuid.uuid4().hex[:12].upper()}"
                
                # Determinar si requiere revisión
                requiere_revision = (
                    extracted_data['confidence'] < 0.7 or
                    not extracted_data['numero_factura'] or
                    extracted_data['monto'] == Decimal('0.00')
                )
                
                invoice = Invoice.objects.create(
                    uploaded_file=uploaded_file,
                    proveedor=proveedor,
                    proveedor_nombre=proveedor.nombre,
                    proveedor_nit=proveedor.nit or '',
                    tipo_proveedor=proveedor.tipo,
                    proveedor_categoria=proveedor.categoria,
                    numero_factura=numero_factura,
                    fecha_emision=extracted_data['fecha_emision'],
                    monto=extracted_data['monto'],
                    tipo_costo=tipo_costo,
                    estado_provision='pendiente',
                    estado_facturacion='pendiente',
                    processing_source='upload_auto' if auto_parse else 'upload_manual',
                    processed_by=str(request.user.id),
                    processed_at=timezone.now(),
                    requiere_revision=requiere_revision,
                    confianza_match=Decimal(str(extracted_data['confidence'])),
                )
                
                # Si hay OT matcheada, asignarla
                if extracted_data['ot_matched']:
                    try:
                        ot = OT.objects.get(numero_ot=extracted_data['ot_matched'])
                        invoice.ot = ot
                        invoice.ot_number = ot.numero_ot
                        invoice.save()
                    except OT.DoesNotExist:
                        pass
                
                # Preparar respuesta
                result_item = {
                    'filename': file.name,
                    'file_id': uploaded_file.id,
                    'invoice_id': invoice.id,
                    'numero_factura': invoice.numero_factura,
                    'monto': float(invoice.monto),
                    'fecha_emision': invoice.fecha_emision.isoformat(),
                    'confidence': float(extracted_data['confidence']),
                    'requiere_revision': requiere_revision,
                    'ot_matched': extracted_data['ot_matched'],
                    'match_method': extracted_data.get('match_method'),  # MBL o Contenedor
                    'numero_contenedor': extracted_data['numero_contenedor'],
                    'mbl': extracted_data['mbl'],
                    'size': uploaded_file.size,
                }
                
                # Mensaje descriptivo
                messages = []
                if auto_parse:
                    if extracted_data['numero_factura']:
                        messages.append(f"✓ Factura detectada: {extracted_data['numero_factura']}")
                    if extracted_data['monto'] and extracted_data['monto'] > 0:
                        messages.append(f"✓ Monto: ${extracted_data['monto']}")
                    if extracted_data['mbl']:
                        messages.append(f"✓ MBL: {extracted_data['mbl']}")
                    if extracted_data['numero_contenedor']:
                        messages.append(f"✓ Contenedor: {extracted_data['numero_contenedor']}")
                    if extracted_data['ot_matched']:
                        match_info = f"✓ OT asignada: {extracted_data['ot_matched']}"
                        if extracted_data.get('match_method'):
                            match_info += f" (por {extracted_data['match_method']})"
                        messages.append(match_info)
                    elif extracted_data['mbl'] or extracted_data['numero_contenedor']:
                        # Tiene MBL/Contenedor pero no encontró OT
                        messages.append("⚠ No se encontró OT para MBL/Contenedor")
                    if requiere_revision:
                        messages.append("⚠ Requiere revisión manual")
                else:
                    messages.append("Factura creada. Debe editarla manualmente.")
                
                result_item['message'] = " | ".join(messages) if messages else "Procesada"
                
                results['success'].append(result_item)
                
            except Exception as e:
                logger.error(f"Error procesando {file.name}: {e}", exc_info=True)
                results['errors'].append({
                    'filename': file.name,
                    'error': str(e)
                })
        
        return Response({
            'total': len(files),
            'processed': len(results['success']),
            'duplicates': len(results['duplicates']),
            'errors': len(results['errors']),
            'results': results,
            'patterns_available': len(results['patterns_used']),
        })
    
    @action(detail=False, methods=['get'])
    def pending(self, request):
        """
        Listar facturas pendientes de revisión.
        Ordenadas por fecha de emisión descendente.
        """
        queryset = self.get_queryset().filter(requiere_revision=True)
        
        # Ordenar
        queryset = queryset.order_by('-fecha_emision', '-created_at')
        
        # Paginar
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = InvoiceListSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        serializer = InvoiceListSerializer(queryset, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def stats(self, request):
        """
        Estadísticas de facturas.
        
        Query params opcionales:
        - fecha_desde: YYYY-MM-DD
        - fecha_hasta: YYYY-MM-DD
        - incluir_excluidas: true/false (por defecto false)
        
        NOTA: Por defecto se excluyen facturas anuladas, rechazadas y disputadas
        de las estadísticas de cuentas por pagar.
        """
        queryset = self.get_queryset()
        
        # Filtrar por rango de fechas si se proporciona
        fecha_desde = request.query_params.get('fecha_desde')
        fecha_hasta = request.query_params.get('fecha_hasta')
        
        if fecha_desde:
            queryset = queryset.filter(fecha_emision__gte=fecha_desde)
        if fecha_hasta:
            queryset = queryset.filter(fecha_emision__lte=fecha_hasta)
        
        # EXCLUIR facturas que no deben contabilizarse (anuladas, rechazadas, disputadas)
        incluir_excluidas = request.query_params.get('incluir_excluidas', 'false').lower() == 'true'
        if not incluir_excluidas:
            queryset = queryset.exclude(estado_provision__in=['anulada', 'rechazada', 'disputada'])
        
        # Calcular estadísticas
        total = queryset.count()
        provisionadas = queryset.filter(estado_provision='provisionada').count()
        disputadas = queryset.filter(estado_provision='disputada').count()
        pendientes_queryset = queryset.filter(
            estado_provision='pendiente',
            fecha_provision__isnull=True
        )
        pendientes_provision = pendientes_queryset.count()
        facturas_sin_fecha_provision = pendientes_provision
        facturadas = queryset.filter(estado_facturacion='facturada').count()
        sin_ot = queryset.filter(ot__isnull=True).count()

        # Estadísticas adicionales de disputas y anulaciones
        total_disputadas = disputadas
        total_anuladas = queryset.filter(estado_provision='anulada').count()
        total_anuladas_parcial = queryset.filter(estado_provision='anulada_parcialmente').count()
        
        # Monto total (usar monto_aplicable si existe, sino monto)
        # Usar Coalesce para obtener monto_aplicable si existe, sino monto
        from django.db.models import F, Case, When
        from django.db.models.functions import Coalesce
        
        total_monto = queryset.aggregate(
            total=Sum(Coalesce('monto_aplicable', 'monto'))
        )['total'] or Decimal('0.00')
        
        # Por tipo de costo
        por_tipo_costo = {}
        for tipo, label in Invoice.TIPO_COSTO_CHOICES:
            count = queryset.filter(tipo_costo=tipo).count()
            monto = queryset.filter(tipo_costo=tipo).aggregate(
                total=Sum(Coalesce('monto_aplicable', 'monto'))
            )['total'] or Decimal('0.00')
            if count > 0:
                por_tipo_costo[label] = {
                    'count': count,
                    'monto': float(monto)
                }
        
        # Top 10 proveedores
        por_proveedor = list(
            queryset.values('proveedor_nombre')
            .annotate(
                count=Count('id'),
                total_monto=Sum(Coalesce('monto_aplicable', 'monto'))
            )
            .order_by('-total_monto')[:10]
        )
        
        data = {
            'total': total,
            'pendientes_revision': disputadas,
            'provisionadas': provisionadas,
            'pendientes_provision': pendientes_provision,
            'sin_fecha_provision': facturas_sin_fecha_provision,
            'facturadas': facturadas,
            'sin_ot': sin_ot,
            'total_monto': total_monto,
            'por_tipo_costo': por_tipo_costo,
            'por_proveedor': por_proveedor,
            # Estadísticas de facturas excluidas
            'total_disputadas': total_disputadas,
            'total_anuladas': total_anuladas,
            'total_anuladas_parcial': total_anuladas_parcial,
        }
        
        serializer = InvoiceStatsSerializer(data)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def assign_ot(self, request, pk=None):
        """
        Asignar OT manualmente a una factura.
        
        Body:
        {
            "ot_id": 123,
            "notas": "Asignación manual por..."
        }
        """
        invoice = self.get_object()
        
        ot_id = request.data.get('ot_id')
        if not ot_id:
            return Response(
                {'error': 'Se requiere ot_id'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        from ots.models import OT
        try:
            ot = OT.objects.get(id=ot_id)
        except OT.DoesNotExist:
            return Response(
                {'error': f'No existe OT con id {ot_id}'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Asignar
        invoice.ot = ot
        invoice.ot_number = ot.numero_ot
        invoice.assignment_method = 'manual'
        invoice.requiere_revision = False
        invoice.confianza_match = Decimal('1.000')
        
        # Agregar nota si se proporciona
        notas = request.data.get('notas')
        if notas:
            timestamp = timezone.now().strftime('%Y-%m-%d %H:%M')
            user = request.user.username
            nueva_nota = f"[{timestamp}] {user}: {notas}"
            if invoice.notas:
                invoice.notas += f"\n{nueva_nota}"
            else:
                invoice.notas = nueva_nota
        
        invoice.save()
        
        serializer = InvoiceDetailSerializer(invoice)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def unassign_ot(self, request, pk=None):
        """Desasignar OT de una factura"""
        invoice = self.get_object()

        invoice.ot = None
        invoice.ot_number = ''
        invoice.assignment_method = ''
        invoice.requiere_revision = True
        invoice.confianza_match = Decimal('0.000')

        # Agregar nota
        notas = request.data.get('notas', 'OT desasignada')
        timestamp = timezone.now().strftime('%Y-%m-%d %H:%M')
        user = request.user.username
        nueva_nota = f"[{timestamp}] {user}: {notas}"
        if invoice.notas:
            invoice.notas += f"\n{nueva_nota}"
        else:
            invoice.notas = nueva_nota

        invoice.save()

        serializer = InvoiceDetailSerializer(invoice)
        return Response(serializer.data)

    @action(detail=False, methods=['post'], url_path='bulk-delete')
    def bulk_delete(self, request):
        """
        Eliminar (soft delete) múltiples facturas.

        Body:
        {
            "invoice_ids": [1, 2, 3, ...]
        }
        """
        invoice_ids = request.data.get('invoice_ids', [])

        if not invoice_ids:
            return Response(
                {'error': 'Debe proporcionar invoice_ids'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Filtrar solo las facturas que existen y no están ya eliminadas
        invoices_to_delete = Invoice.objects.filter(id__in=invoice_ids, is_deleted=False)
        deleted_count = invoices_to_delete.update(is_deleted=True)

        return Response(
            {'message': f'{deleted_count} facturas eliminadas exitosamente.'},
            status=status.HTTP_200_OK
        )

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        """
        Exportar facturas a Excel con formato profesional.
        Respeta todos los filtros aplicados en get_queryset().
        EXPORTA TODOS LOS REGISTROS FILTRADOS (no solo la página actual).
        
        Retorna un archivo Excel con:
        - Fechas en formato dd/mm/yyyy
        - Montos con formato contable ($)
        - Encabezados con estilo profesional
        - Anchos de columna ajustados
        - Colores y formato de tabla
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # Forzar la no paginación estableciendo un tamaño de página muy grande
        if self.paginator:
            self.paginator.page_size = 10000
        queryset = self.filter_queryset(self.get_queryset())

        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Facturas"

        # Estilos profesionales
        header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")  # Azul oscuro
        header_font = Font(color="FFFFFF", bold=True, size=12, name='Calibri')
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Estilo para filas alternas
        alt_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")  # Gris claro
        
        # Bordes
        thin_border = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='thin', color='D1D5DB')
        )
        
        # Fuente para datos
        data_font = Font(size=11, name='Calibri')
        data_alignment = Alignment(horizontal="left", vertical="center")

        headers = [
            'ID', 'Número Factura', 'OT', 'Cliente', 'MBL', 'Contenedor',
            'Naviera', 'Barco', 'Proveedor', 'NIT Proveedor', 'Tipo Proveedor',
            'Tipo Costo', 'Monto (USD)', 'Fecha Emisión', 'Fecha Vencimiento',
            'Fecha Provisión', 'Fecha Facturación', 'Estado Provisión',
            'Estado Facturación', 'Método Asignación', 'Confianza Match',
            'Requiere Revisión', 'Notas', 'Creado'
        ]

        # Escribir headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        # Escribir datos
        try:
            for row_num, invoice in enumerate(queryset, 2):
                # Obtener datos relacionados
                ot_number = invoice.ot.numero_ot if invoice.ot else ''
                cliente = invoice.ot.cliente.normalized_name if invoice.ot and invoice.ot.cliente else ''
                mbl = invoice.ot.master_bl if invoice.ot else ''
                contenedor = ', '.join(invoice.ot.contenedores) if invoice.ot and invoice.ot.contenedores else ''
                naviera = invoice.ot.proveedor.nombre if invoice.ot and invoice.ot.proveedor else ''
                barco = invoice.ot.barco if invoice.ot else ''

                row_data = [
                    invoice.id,
                    invoice.numero_factura or '',
                    ot_number,
                    cliente,
                    mbl,
                    contenedor,
                    naviera,
                    barco,
                    invoice.proveedor.nombre if invoice.proveedor else invoice.proveedor_nombre,
                    invoice.proveedor_nit or '',
                    invoice.get_tipo_proveedor_display() if invoice.tipo_proveedor else '',
                    invoice.get_tipo_costo_display() if invoice.tipo_costo else '',
                    float(invoice.monto) if invoice.monto else 0.0,
                    invoice.fecha_emision,
                    invoice.fecha_vencimiento,
                    invoice.fecha_provision,
                    invoice.fecha_facturacion,
                    invoice.get_estado_provision_display() if invoice.estado_provision else '',
                    invoice.get_estado_facturacion_display() if invoice.estado_facturacion else '',
                    invoice.assignment_method or '',
                    float(invoice.confianza_match) if invoice.confianza_match else 0.0,
                    'Sí' if invoice.requiere_revision else 'No',
                    invoice.notas or '',
                    invoice.created_at,
                ]

                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.border = thin_border
                    cell.font = data_font
                    cell.alignment = data_alignment
                    
                    # Aplicar fondo alterno para mejor legibilidad
                    if row_num % 2 == 0:
                        cell.fill = alt_fill

                    # Formato de fechas (DD/MM/YYYY) - columnas 14, 15, 16, 17, 24
                    if col_num in [14, 15, 16, 17, 24]:
                        if value:
                            cell.number_format = 'DD/MM/YYYY'
                            cell.alignment = Alignment(horizontal="center", vertical="center")

                    # Formato de moneda contable - columna 13 (Monto USD)
                    if col_num == 13:
                        cell.number_format = '$#,##0.00'
                        cell.alignment = Alignment(horizontal="right", vertical="center")

                    # Formato de porcentaje - columna 21 (Confianza Match)
                    if col_num == 21:
                        cell.number_format = '0.0%'
                        cell.alignment = Alignment(horizontal="center", vertical="center")
        except Exception as e:
            logger = logging.getLogger(__name__)
            logger.error(f"Error al exportar facturas a Excel: {e}", exc_info=True)
            return Response(
                {"error": "Ocurrió un error inesperado al generar el archivo Excel."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        # Ajustar anchos de columna
        column_widths = [
            8, 20, 15, 25, 18, 15, 20, 20, 25, 15, 18, 15, 12,
            15, 18, 18, 20, 18, 20, 20, 15, 15, 40, 20
        ]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width

        # Congelar primera fila
        ws.freeze_panes = 'A2'
        
        # Agregar filtros de Excel (autofilter)
        ws.auto_filter.ref = ws.dimensions

        # Crear respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        total_records = queryset.count()
        filename = f'Facturas_{total_records}_registros_{timestamp}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response['Access-Control-Expose-Headers'] = 'Content-Disposition'
        wb.save(response)

        return response

    @action(detail=False, methods=['post'], url_path='bulk-pdf')
    def bulk_pdf(self, request):
        """
        Descargar múltiples facturas como PDFs individuales.
        Si es una sola factura, descarga el PDF directamente.
        Si son múltiples, las empaqueta en un ZIP.
        
        Formato de nombres: FACTURA PROVEEDOR NUMERO ALIAS OT.pdf

        Body:
        {
            "invoice_ids": [1, 2, 3, ...]
        }
        """
        logger = logging.getLogger(__name__)

        invoice_ids = request.data.get('invoice_ids', [])
        if not invoice_ids:
            return Response(
                {'error': 'Debe proporcionar invoice_ids'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Obtener facturas
        invoices = Invoice.objects.filter(
            id__in=invoice_ids, 
            is_deleted=False
        ).select_related('ot', 'ot__cliente', 'proveedor', 'uploaded_file')

        if not invoices.exists():
            return Response(
                {'error': 'No se encontraron facturas'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Si es solo una factura, descargarla directamente
        if len(invoice_ids) == 1:
            invoice = invoices.first()
            
            if not invoice.uploaded_file:
                return Response(
                    {'error': 'La factura no tiene archivo asociado'},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            storage_path = invoice.uploaded_file.path
            if not get_storage().exists(storage_path):
                return Response(
                    {'error': 'Archivo no encontrado'},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            try:
                file_handle = get_storage().open(storage_path, 'rb')
                filename = self._generate_friendly_filename(invoice)
                if not filename:
                    filename = invoice.uploaded_file.filename or 'factura.pdf'
                
                content_type = invoice.uploaded_file.content_type or 'application/pdf'
                response = FileResponse(file_handle, content_type=content_type)
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                response['Content-Length'] = invoice.uploaded_file.size
                response['Access-Control-Expose-Headers'] = 'Content-Disposition'
                
                return response
            except Exception as e:
                logger.error(f"Error descargando factura {invoice.id}: {e}")
                return Response(
                    {'error': 'Error al descargar el archivo'},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )

        # Si son múltiples facturas, crear ZIP
        zip_buffer = BytesIO()
        processed_count = 0

        try:
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                for invoice in invoices:
                    if not invoice.uploaded_file:
                        logger.warning(f"Factura {invoice.id} no tiene archivo asociado")
                        continue

                    storage_path = invoice.uploaded_file.path

                    if not get_storage().exists(storage_path):
                        logger.warning(f"Archivo no encontrado: {storage_path}")
                        continue

                    try:
                        # Leer archivo del storage
                        with get_storage().open(storage_path, 'rb') as file_handle:
                            file_content = file_handle.read()

                        # Generar nombre amigable
                        filename = self._generate_friendly_filename(invoice)
                        if not filename:
                            filename = invoice.uploaded_file.filename or f'factura_{invoice.id}.pdf'

                        # Agregar al ZIP
                        zip_file.writestr(filename, file_content)
                        processed_count += 1

                    except Exception as e:
                        logger.error(f"Error procesando factura {invoice.id}: {e}")
                        continue

            if processed_count == 0:
                return Response(
                    {'error': 'No se pudo procesar ninguna factura'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Preparar respuesta ZIP
            zip_buffer.seek(0)
            zip_content = zip_buffer.read()
            
            response = HttpResponse(zip_content, content_type='application/zip')
            timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
            response['Content-Disposition'] = f'attachment; filename="Facturas_PDF_{timestamp}.zip"'
            response['Access-Control-Expose-Headers'] = 'Content-Disposition'
            response['Content-Length'] = len(zip_content)

            return response
            
        except Exception as e:
            logger.error(f"Error creando ZIP: {e}")
            return Response(
                {'error': f'Error al crear archivo ZIP: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'], url_path='bulk-zip')
    def bulk_zip(self, request):
        """
        Exportar facturas seleccionadas en ZIP con estructura de carpetas:
        CLIENTE/OT/FACTURA PROVEEDOR NUMERO_FACTURA ALIAS_CLIENTE OT.pdf
        (con espacios, sin guiones)

        Body:
        {
            "invoice_ids": [1, 2, 3, ...]
        }
        """
        logger = logging.getLogger(__name__)

        invoice_ids = request.data.get('invoice_ids', [])
        if not invoice_ids:
            return Response(
                {'error': 'Debe proporcionar invoice_ids'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Obtener facturas
        invoices = Invoice.objects.filter(
            id__in=invoice_ids,
            is_deleted=False
        ).select_related('ot', 'ot__cliente', 'proveedor', 'uploaded_file')

        if not invoices.exists():
            return Response(
                {'error': 'No se encontraron facturas'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Crear ZIP en memoria
        zip_buffer = BytesIO()
        processed_count = 0
        skipped_no_ot = 0

        try:
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                for invoice in invoices:
                    if not invoice.uploaded_file:
                        logger.warning(f"Factura {invoice.id} no tiene archivo asociado")
                        continue

                    if not invoice.ot:
                        logger.warning(f"Factura {invoice.id} no tiene OT asignada, se omitirá")
                        skipped_no_ot += 1
                        continue

                    storage_path = invoice.uploaded_file.path

                    if not get_storage().exists(storage_path):
                        logger.warning(f"Archivo no encontrado: {storage_path}")
                        continue

                    try:
                        # Leer archivo del storage
                        with get_storage().open(storage_path, 'rb') as file_handle:
                            file_content = file_handle.read()

                        # 1. Carpeta de Cliente (usar short_name o normalized_name)
                        if invoice.ot.cliente:
                            cliente_name = invoice.ot.cliente.short_name or invoice.ot.cliente.normalized_name
                            # Limpiar y mantener espacios en carpetas
                            cliente_folder = re.sub(r'[^\w\s]', '', cliente_name).strip()[:50]
                        else:
                            cliente_folder = 'SIN CLIENTE'

                        # 2. Carpeta de OT (limpiar caracteres especiales)
                        ot_folder = re.sub(r'[^\w]', '', invoice.ot.numero_ot)[:50]

                        # 3. Nombre del archivo
                        proveedor_name = invoice.proveedor.nombre if invoice.proveedor else invoice.proveedor_nombre
                        proveedor_name = re.sub(r'[^\w\s]', '', proveedor_name).strip()[:30]

                        numero_factura = re.sub(r'[^\w]', '', invoice.numero_factura)[:30]
                        
                        cliente_short = ''
                        if invoice.ot.cliente and invoice.ot.cliente.short_name:
                            cliente_short = re.sub(r'[^\w\s]', '', invoice.ot.cliente.short_name).strip()[:20]
                        
                        ot_number = re.sub(r'[^\w]', '', invoice.ot.numero_ot)[:20]

                        # Obtener extensión
                        original_filename = invoice.uploaded_file.filename or 'invoice.pdf'
                        _, ext = os.path.splitext(original_filename)
                        if not ext:
                            ext = '.pdf'

                        # Construir nombre con espacios (sin guiones)
                        parts = ['FACTURA', proveedor_name, numero_factura]
                        if cliente_short:
                            parts.append(cliente_short)
                        if ot_number:
                            parts.append(ot_number)
                        
                        filename = ' '.join(filter(None, parts)) + ext

                        # Si es muy largo (>150 chars), usar formato corto
                        if len(filename) > 150:
                            filename = f"{numero_factura} {ot_number}{ext}"

                        # Ruta completa en el ZIP (usar / para compatibilidad)
                        zip_path = f"{cliente_folder}/{ot_folder}/{filename}"

                        # Agregar al ZIP
                        zip_file.writestr(zip_path, file_content)
                        processed_count += 1

                    except Exception as e:
                        logger.error(f"Error procesando factura {invoice.id}: {e}")
                        continue

            if processed_count == 0:
                error_msg = 'No se pudo procesar ninguna factura'
                if skipped_no_ot > 0:
                    error_msg += f'. {skipped_no_ot} factura(s) sin OT asignada'
                return Response(
                    {'error': error_msg},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Preparar respuesta (fuera del context manager)
            zip_buffer.seek(0)
            zip_content = zip_buffer.read()
            
            response = HttpResponse(zip_content, content_type='application/zip')
            timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
            response['Content-Disposition'] = f'attachment; filename="Facturas_Estructuradas_{timestamp}.zip"'
            response['Access-Control-Expose-Headers'] = 'Content-Disposition'
            response['Content-Length'] = len(zip_content)

            return response

        except Exception as e:
            logger.error(f"Error creando ZIP estructurado: {e}")
            return Response(
                {'error': f'Error al crear archivo ZIP: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class UploadedFileViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet de solo lectura para archivos subidos.
    Útil para verificar duplicados y gestionar archivos.
    """

    permission_classes = [IsAuthenticated]
    queryset = UploadedFile.objects.all()
    serializer_class = UploadedFileSerializer

    def get_queryset(self):
        """Filtrar por hash si se proporciona"""
        queryset = super().get_queryset()

        sha256 = self.request.query_params.get('sha256')
        if sha256:
            queryset = queryset.filter(sha256=sha256)

        return queryset.order_by('-created_at')


class DisputeViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestión de disputas.

    Endpoints:
    - GET /disputes/ - Listar disputas
    - POST /disputes/ - Crear disputa
    - GET /disputes/{id}/ - Detalle de disputa
    - PUT/PATCH /disputes/{id}/ - Actualizar disputa
    - DELETE /disputes/{id}/ - Eliminar (soft delete) disputa
    - GET /disputes/stats/ - Estadísticas
    """

    queryset = Dispute.objects.filter(is_deleted=False).select_related(
        'invoice', 'ot', 'invoice__proveedor', 'ot__cliente'
    )
    permission_classes = [IsAuthenticated]
    http_method_names = ['get', 'post', 'put', 'patch', 'delete', 'head', 'options']

    def get_serializer_class(self):
        if self.action == 'list':
            return DisputeListSerializer
        elif self.action == 'create':
            return DisputeCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return DisputeUpdateSerializer
        return DisputeDetailSerializer

    def get_queryset(self):
        """
        Filtrado de disputas.
        Query params:
        - estado: abierta, en_revision, resuelta, cerrada
        - tipo_disputa: cantidad, servicio, duplicada, precio, otro
        - invoice_id: id de factura
        - ot_id: id de OT
        - search: búsqueda general
        """
        queryset = super().get_queryset()

        # Filtros
        estado = self.request.query_params.get('estado')
        if estado:
            queryset = queryset.filter(estado=estado)

        tipo_disputa = self.request.query_params.get('tipo_disputa')
        if tipo_disputa:
            queryset = queryset.filter(tipo_disputa=tipo_disputa)

        invoice_id = self.request.query_params.get('invoice_id')
        if invoice_id:
            queryset = queryset.filter(invoice_id=invoice_id)

        ot_id = self.request.query_params.get('ot_id')
        if ot_id:
            queryset = queryset.filter(ot_id=ot_id)

        # Filtro de resultado
        resultado = self.request.query_params.get('resultado')
        if resultado:
            queryset = queryset.filter(resultado=resultado)

        # Búsqueda general
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(numero_caso__icontains=search) |
                Q(detalle__icontains=search) |
                Q(invoice__numero_factura__icontains=search) |
                Q(ot__numero_ot__icontains=search) |
                Q(operativo__icontains=search) |
                Q(ot__operativo__icontains=search)
            )

        return queryset.order_by('-created_at')

    def create(self, request, *args, **kwargs):
        """
        Crear una nueva disputa.
        """
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)
    
    @action(detail=True, methods=['get'])
    def eventos(self, request, pk=None):
        """
        Obtener timeline de eventos de una disputa.
        """
        dispute = self.get_object()
        eventos = dispute.eventos.all()
        serializer = DisputeEventSerializer(eventos, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def add_evento(self, request, pk=None):
        """
        Agregar un nuevo evento/comentario a la disputa.
        """
        dispute = self.get_object()
        serializer = DisputeEventSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(
                dispute=dispute,
                usuario=request.user.username if request.user else ''
            )
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'])
    def resolve(self, request, pk=None):
        """
        Resolver disputa con opción de crear nota de crédito.

        Este endpoint permite:
        1. Resolver una disputa (estado, resultado, monto_recuperado)
        2. Opcionalmente crear una nota de crédito asociada

        Body:
        {
            "estado": "resuelta",
            "resultado": "aprobada_total" | "aprobada_parcial" | "rechazada",
            "monto_recuperado": 5000.00,  // requerido para aprobada_parcial
            "resolucion": "Descripción de la resolución",

            // Campos opcionales para nota de crédito
            "tiene_nota_credito": true,
            "nota_credito_numero": "NC-2024-001",
            "nota_credito_monto": 5000.00,
            "nota_credito_archivo": <file>  // multipart/form-data
        }
        """
        from .serializers import DisputeResolveSerializer

        dispute = self.get_object()

        # Validar datos
        serializer = DisputeResolveSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        # Actualizar disputa
        dispute.estado = serializer.validated_data['estado']
        dispute.resultado = serializer.validated_data['resultado']
        dispute.monto_recuperado = serializer.validated_data.get('monto_recuperado', Decimal('0.00'))
        dispute.resolucion = serializer.validated_data.get('resolucion', '')

        # Si es aprobada_total, auto-asignar monto_disputa como monto_recuperado
        if dispute.resultado == 'aprobada_total':
            dispute.monto_recuperado = dispute.monto_disputa

        dispute.save()

        # Crear nota de crédito si aplica
        if serializer.validated_data.get('tiene_nota_credito'):
            uploaded_file = None
            nota_credito_archivo = request.FILES.get('nota_credito_archivo')

            # Procesar archivo si se proporcionó
            if nota_credito_archivo:
                # Calcular hash
                nota_credito_archivo.seek(0)
                file_content = nota_credito_archivo.read()
                file_hash = UploadedFile.calculate_hash(file_content)

                # Verificar si ya existe
                existing_file = UploadedFile.objects.filter(sha256=file_hash).first()

                if existing_file:
                    uploaded_file = existing_file
                else:
                    # Guardar archivo nuevo
                    nota_credito_archivo.seek(0)
                    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
                    safe_filename = f"{timestamp}_{nota_credito_archivo.name}"
                    path = get_storage().save(f'credit_notes/{safe_filename}', nota_credito_archivo)

                    uploaded_file = UploadedFile.objects.create(
                        filename=nota_credito_archivo.name,
                        path=path,
                        sha256=file_hash,
                        size=nota_credito_archivo.size,
                        content_type=nota_credito_archivo.content_type
                    )

            # Crear nota de crédito
            monto_nc = serializer.validated_data['nota_credito_monto']
            CreditNote.objects.create(
                numero_nota=serializer.validated_data['nota_credito_numero'],
                invoice_relacionada=dispute.invoice,
                proveedor=dispute.invoice.proveedor,
                proveedor_nombre=dispute.invoice.proveedor_nombre,
                fecha_emision=timezone.localdate(),
                monto=-abs(monto_nc),  # Asegurar que sea negativo
                motivo=f'Nota de crédito por disputa {dispute.numero_caso} - {dispute.get_resultado_display()}',
                estado='aplicada',  # Aplicar automáticamente
                uploaded_file=uploaded_file,
                processed_by=request.user.username if request.user else 'system',
                processed_at=timezone.now(),
                processing_source='manual_entry'
            )

        # Crear evento de resolución
        DisputeEvent.objects.create(
            dispute=dispute,
            tipo='resolucion',
            descripcion=f'Disputa resuelta: {dispute.get_resultado_display()}. {dispute.resolucion}',
            usuario=request.user.username if request.user else '',
            monto_recuperado=dispute.monto_recuperado if dispute.monto_recuperado > 0 else None
        )

        # Retornar disputa actualizada
        from .serializers import DisputeDetailSerializer
        return Response(DisputeDetailSerializer(dispute).data)

    @action(detail=False, methods=['get'])
    def stats(self, request):
        """
        Estadísticas de disputas.
        """
        queryset = self.get_queryset()

        # Calcular estadísticas
        total = queryset.count()
        abiertas = queryset.filter(estado='abierta').count()
        en_revision = queryset.filter(estado='en_revision').count()
        resueltas = queryset.filter(estado='resuelta').count()
        cerradas = queryset.filter(estado='cerrada').count()

        # Monto total en disputa
        total_monto = queryset.filter(
            estado__in=['abierta', 'en_revision']
        ).aggregate(total=Sum('monto_disputa'))['total'] or Decimal('0.00')

        # Por tipo de disputa
        por_tipo = {}
        for tipo, label in Dispute.TIPO_DISPUTA_CHOICES:
            count = queryset.filter(tipo_disputa=tipo).count()
            if count > 0:
                por_tipo[label] = count

        data = {
            'total': total,
            'abiertas': abiertas,
            'en_revision': en_revision,
            'resueltas': resueltas,
            'cerradas': cerradas,
            'total_monto_disputado': float(total_monto),
            'por_tipo': por_tipo,
        }

        return Response(data)
    
    @action(detail=False, methods=['get'])
    def filter_values(self, request):
        """
        Obtener valores únicos para los filtros.
        """
        # Usar queryset base sin filtros aplicados para obtener todos los valores posibles
        base_queryset = Dispute.objects.filter(is_deleted=False)
        
        # Obtener valores únicos de cada campo usando set para garantizar unicidad
        estados = set(base_queryset.values_list('estado', flat=True))
        tipos_disputa = set(base_queryset.values_list('tipo_disputa', flat=True))
        resultados = set(base_queryset.exclude(resultado__isnull=True).values_list('resultado', flat=True))
        
        # Filtrar valores nulos y ordenar
        data = {
            'estados': sorted([e for e in estados if e]),
            'tipos_disputa': sorted([t for t in tipos_disputa if t]),
            'resultados': sorted([r for r in resultados if r]),
        }
        
        return Response(data)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_dispute(request):
    """
    Endpoint para crear disputas.
    """
    serializer = DisputeCreateSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class CreditNoteViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestión de notas de crédito.
    """

    permission_classes = [IsAuthenticated]
    queryset = CreditNote.objects.filter(is_deleted=False).select_related(
        'proveedor', 'invoice_relacionada', 'uploaded_file'
    )

    def get_serializer_class(self):
        if self.action == 'list':
            return CreditNoteListSerializer
        elif self.action == 'create':
            return CreditNoteCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return CreditNoteUpdateSerializer
        return CreditNoteDetailSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        estado = self.request.query_params.get('estado')
        if estado:
            queryset = queryset.filter(estado=estado)
        proveedor_id = self.request.query_params.get('proveedor_id')
        if proveedor_id:
            queryset = queryset.filter(proveedor_id=proveedor_id)
        invoice_id = self.request.query_params.get('invoice_id')
        if invoice_id:
            queryset = queryset.filter(invoice_relacionada_id=invoice_id)
        fecha_desde = self.request.query_params.get('fecha_desde')
        if fecha_desde:
            queryset = queryset.filter(fecha_emision__gte=fecha_desde)
        fecha_hasta = self.request.query_params.get('fecha_hasta')
        if fecha_hasta:
            queryset = queryset.filter(fecha_emision__lte=fecha_hasta)
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(numero_nota__icontains=search) |
                Q(proveedor_nombre__icontains=search) |
                Q(motivo__icontains=search)
            )
        return queryset.order_by('-fecha_emision', '-created_at')

    def create(self, request, *args, **kwargs):
        """
        Crea una nota de crédito manualmente desde el modal.
        Extrae el proveedor de la factura relacionada.
        """
        invoice_id = request.data.get('invoice_relacionada_id')
        if not invoice_id:
            return Response(
                {'invoice_relacionada_id': ['Este campo es requerido.']},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            invoice = Invoice.objects.get(id=invoice_id)
        except Invoice.DoesNotExist:
            return Response(
                {'invoice_relacionada_id': ['La factura especificada no existe.']},
                status=status.HTTP_404_NOT_FOUND
            )

        # Validar que la factura no esté ya saldada
        if invoice.get_monto_aplicable() <= 0:
            return Response(
                {'invoice_relacionada_id': ['Esta factura ya ha sido saldada o anulada y no admite más notas de crédito.']},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Preparar datos para el serializador
        data = request.data.copy()
        data['proveedor_id'] = invoice.proveedor.id
        data['proveedor_nombre'] = invoice.proveedor.nombre
        
        # Mapear el archivo del frontend al campo del serializador
        pdf_file = request.FILES.get('pdf_file')
        if pdf_file:
            data['file'] = pdf_file

        serializer = CreditNoteCreateSerializer(data=data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        
        # El método create del serializador se encargará de guardar
        credit_note = serializer.save(estado='aplicada')

        # Devolver la respuesta con el serializador de detalle
        return Response(
            CreditNoteDetailSerializer(credit_note).data,
            status=status.HTTP_201_CREATED
        )

    @action(detail=False, methods=['post'], url_path='upload')
    def upload(self, request):
        from .parsers.pdf_extractor import PDFExtractor
        from .parsers.pattern_service import PatternApplicationService
        
        logger = logging.getLogger(__name__)
        
        files = request.FILES.getlist('files[]')
        proveedor_id = request.data.get('proveedor_id')
        auto_parse = request.data.get('auto_parse', 'true').lower() == 'true'
        
        if not files:
            return Response({'error': 'No se enviaron archivos'}, status=status.HTTP_400_BAD_REQUEST)
        if not proveedor_id:
            return Response({'error': 'Debe seleccionar un proveedor'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            proveedor = Provider.objects.get(id=proveedor_id)
        except Provider.DoesNotExist:
            return Response({'error': 'Proveedor no encontrado'}, status=status.HTTP_400_BAD_REQUEST)
        
        pdf_extractor = PDFExtractor()
        pattern_service = PatternApplicationService(provider_id=int(proveedor_id))
        
        results = {'success': [], 'errors': [], 'duplicates': []}
        
        for file in files:
            try:
                file.seek(0)
                file_content = file.read()
                file_hash = UploadedFile.calculate_hash(file_content)
                
                existing_file = UploadedFile.objects.filter(sha256=file_hash).first()
                if existing_file and CreditNote.objects.filter(uploaded_file=existing_file, is_deleted=False).exists():
                    results['duplicates'].append({'filename': file.name, 'reason': f'Archivo duplicado con nota de crédito activa'})
                    continue

                if not existing_file:
                    file.seek(0)
                    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
                    safe_filename = f"{timestamp}_{file.name}"
                    path = get_storage().save(f'credit_notes/{safe_filename}', file)
                    uploaded_file = UploadedFile.objects.create(
                        filename=file.name, path=path, sha256=file_hash, size=file.size, content_type=file.content_type
                    )
                else:
                    uploaded_file = existing_file
                
                extracted_data = {
                    'numero_nota': None, 'fecha_emision': timezone.localdate(),
                    'monto': Decimal('0.00'), 'invoice_relacionada': None
                }
                
                if auto_parse and file.content_type == 'application/pdf':
                    try:
                        file.seek(0)
                        text = pdf_extractor.extract(file.read()).text
                        if text:
                            pattern_results = pattern_service.apply_patterns(text)
                            if 'numero_nota_credito' in pattern_results:
                                extracted_data['numero_nota'] = pattern_results['numero_nota_credito']['value']
                            if 'monto_total' in pattern_results:
                                extracted_data['monto'] = pattern_results['monto_total']['value']
                            if 'fecha_emision' in pattern_results:
                                fecha = pattern_results['fecha_emision']['value']
                                if isinstance(fecha, datetime):
                                    extracted_data['fecha_emision'] = fecha.date()
                                elif isinstance(fecha, date):
                                    extracted_data['fecha_emision'] = fecha
                            if 'numero_factura' in pattern_results:
                                numero_factura_relacionada = pattern_results['numero_factura']['value']
                                invoice_relacionada = Invoice.objects.filter(numero_factura=numero_factura_relacionada).first()
                                if invoice_relacionada:
                                    extracted_data['invoice_relacionada'] = invoice_relacionada
                    except Exception as e:
                        logger.error(f"Error en auto-parsing para {file.name}: {e}")

                numero_nota = extracted_data['numero_nota'] or f"TEMP-NC-{uuid.uuid4().hex[:10].upper()}"

                credit_note = CreditNote.objects.create(
                    uploaded_file=uploaded_file, proveedor=proveedor, proveedor_nombre=proveedor.nombre,
                    numero_nota=numero_nota, fecha_emision=extracted_data['fecha_emision'],
                    monto=extracted_data['monto'], invoice_relacionada=extracted_data['invoice_relacionada'],
                    motivo="Carga automática", estado='pendiente', processing_source='upload_auto',
                    processed_by=str(request.user.id), processed_at=timezone.now(),
                )
                
                results['success'].append({
                    'filename': file.name, 'credit_note_id': credit_note.id,
                    'numero_nota': credit_note.numero_nota, 'monto': float(credit_note.monto),
                })
                
            except Exception as e:
                logger.error(f"Error procesando {file.name}: {e}", exc_info=True)
                results['errors'].append({'filename': file.name, 'error': str(e)})
        
        return Response({
            'total': len(files), 'processed': len(results['success']),
            'duplicates': len(results['duplicates']), 'errors': len(results['errors']),
            'results': results,
        })

    @action(detail=False, methods=['post'], url_path='manual')
    def manual(self, request):
        """
        Crear nota de crédito manualmente.

        Permite asociar nota de crédito a una factura, auto-completa OT si la factura la tiene.
        No requiere validaciones complejas - es un flujo manual supervisado.

        Body (multipart/form-data):
        {
            "numero_nota": "NC-2024-001",
            "invoice_id": 123,
            "monto": 5000.00,
            "motivo": "Descripción",
            "fecha_emision": "2025-01-15",
            "file": <archivo PDF opcional>
        }
        """
        logger = logging.getLogger(__name__)

        # Validar campos requeridos
        numero_nota = request.data.get('numero_nota')
        invoice_id = request.data.get('invoice_id')
        monto = request.data.get('monto')
        motivo = request.data.get('motivo')
        fecha_emision = request.data.get('fecha_emision', timezone.localdate())

        if not numero_nota:
            return Response({'error': 'El número de nota es obligatorio'}, status=status.HTTP_400_BAD_REQUEST)
        if not invoice_id:
            return Response({'error': 'Debe seleccionar una factura'}, status=status.HTTP_400_BAD_REQUEST)
        if not monto:
            return Response({'error': 'El monto es obligatorio'}, status=status.HTTP_400_BAD_REQUEST)
        if not motivo:
            return Response({'error': 'El motivo es obligatorio'}, status=status.HTTP_400_BAD_REQUEST)

        # Obtener factura
        try:
            invoice = Invoice.objects.get(id=invoice_id, is_deleted=False)
        except Invoice.DoesNotExist:
            return Response({'error': 'Factura no encontrada'}, status=status.HTTP_404_NOT_FOUND)

        # Procesar archivo si se proporcionó
        uploaded_file = None
        file = request.FILES.get('file')
        if file:
            try:
                # Calcular hash
                file.seek(0)
                file_content = file.read()
                file_hash = UploadedFile.calculate_hash(file_content)

                # Verificar si ya existe
                existing_file = UploadedFile.objects.filter(sha256=file_hash).first()

                if existing_file:
                    uploaded_file = existing_file
                else:
                    # Guardar archivo nuevo
                    file.seek(0)
                    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
                    safe_filename = f"{timestamp}_{file.name}"
                    path = get_storage().save(f'credit_notes/{safe_filename}', file)

                    uploaded_file = UploadedFile.objects.create(
                        filename=file.name,
                        path=path,
                        sha256=file_hash,
                        size=file.size,
                        content_type=file.content_type
                    )
            except Exception as e:
                logger.error(f"Error procesando archivo: {e}")
                return Response({'error': 'Error al procesar el archivo'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # Crear nota de crédito
        try:
            monto_decimal = Decimal(str(monto))
            credit_note = CreditNote.objects.create(
                numero_nota=numero_nota,
                invoice_relacionada=invoice,
                proveedor=invoice.proveedor,
                proveedor_nombre=invoice.proveedor_nombre,
                fecha_emision=fecha_emision,
                monto=-abs(monto_decimal),  # Asegurar que sea negativo
                motivo=motivo,
                estado='aplicada',  # Aplicar automáticamente en entrada manual
                uploaded_file=uploaded_file,
                processed_by=request.user.username if request.user else 'system',
                processed_at=timezone.now(),
                processing_source='manual_entry'
            )

            from .serializers import CreditNoteDetailSerializer
            return Response(CreditNoteDetailSerializer(credit_note).data, status=status.HTTP_201_CREATED)

        except Exception as e:
            logger.error(f"Error creando nota de crédito: {e}", exc_info=True)
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['get'], url_path='file')
    def retrieve_file(self, request, pk=None):
        from django.shortcuts import redirect
        from django.conf import settings

        credit_note = self.get_object()
        if not credit_note.uploaded_file:
            return Response({'detail': 'La nota de crédito no tiene archivo asociado.'}, status=status.HTTP_404_NOT_FOUND)

        storage_path = credit_note.uploaded_file.path

        # If using Cloudinary, redirect directly to Cloudinary URL (fast!)
        if getattr(settings, 'USE_CLOUDINARY', False):
            try:
                storage = get_storage()
                cloudinary_url = storage.url(storage_path)
                return redirect(cloudinary_url)
            except Exception as exc:
                return Response(
                    {'detail': f'Error al obtener URL del archivo: {exc}'},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )

        # For local filesystem, serve file normally
        if not get_storage().exists(storage_path):
            return Response({'detail': 'Archivo no encontrado en el almacenamiento.'}, status=status.HTTP_404_NOT_FOUND)
        try:
            file_handle = get_storage().open(storage_path, 'rb')
        except Exception as exc:
            return Response({'detail': f'No se pudo abrir el archivo: {exc}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        filename = credit_note.uploaded_file.filename or f"NC_{credit_note.numero_nota}.pdf"
        content_type = credit_note.uploaded_file.content_type or 'application/octet-stream'
        response = FileResponse(file_handle, content_type=content_type)
        download_flag = str(request.query_params.get('download', '')).lower()
        disposition = 'attachment' if download_flag in ('1', 'true', 'yes') else 'inline'
        response['Content-Disposition'] = f'{disposition}; filename="{filename}"'
        response['Content-Length'] = credit_note.uploaded_file.size
        response['Access-Control-Expose-Headers'] = 'Content-Disposition'
        return response

    @action(detail=False, methods=['get'])
    def stats(self, request):
        """
        Estadísticas de notas de crédito.

        Query params opcionales:
        - fecha_desde: YYYY-MM-DD
        - fecha_hasta: YYYY-MM-DD
        - proveedor_id: filtrar por proveedor
        """
        queryset = self.get_queryset()

        # Filtros opcionales
        fecha_desde = request.query_params.get('fecha_desde')
        if fecha_desde:
            queryset = queryset.filter(fecha_emision__gte=fecha_desde)

        fecha_hasta = request.query_params.get('fecha_hasta')
        if fecha_hasta:
            queryset = queryset.filter(fecha_emision__lte=fecha_hasta)

        proveedor_id = request.query_params.get('proveedor_id')
        if proveedor_id:
            queryset = queryset.filter(proveedor_id=proveedor_id)

        # Contadores por estado
        total_notas = queryset.count()
        pendientes = queryset.filter(estado='pendiente').count()
        aplicadas = queryset.filter(estado='aplicada').count()
        rechazadas = queryset.filter(estado='rechazada').count()

        # Montos (recordar que son negativos)
        monto_total = abs(queryset.aggregate(total=Sum('monto'))['total'] or Decimal('0.00'))
        monto_aplicadas = abs(queryset.filter(estado='aplicada').aggregate(total=Sum('monto'))['total'] or Decimal('0.00'))
        monto_pendientes = abs(queryset.filter(estado='pendiente').aggregate(total=Sum('monto'))['total'] or Decimal('0.00'))

        # Por proveedor (top 10)
        por_proveedor = list(
            queryset.values('proveedor_nombre')
            .annotate(
                count=Count('id'),
                total_monto=Sum('monto')
            )
            .order_by('total_monto')[:10]  # Ordenar ascendente porque son negativos
        )

        data = {
            'total_notas': total_notas,
            'pendientes': pendientes,
            'aplicadas': aplicadas,
            'rechazadas': rechazadas,
            'monto_total': float(monto_total),
            'monto_aplicadas': float(monto_aplicadas),
            'monto_pendientes': float(monto_pendientes),
            'por_proveedor': por_proveedor,
        }
        return Response(data)
