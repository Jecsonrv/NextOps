"""
Views para gestión de Órdenes de Trabajo (OTs).
"""

import re

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters
from django.db.models import Q, Count, Sum

from .models import OT
from .serializers import (
    OTListSerializer,
    OTDetailSerializer,
    OTSearchSerializer,
    ContenedorSerializer,
    ProvisionHierarchySerializer,
    ExcelUploadSerializer,
    ExcelImportResultSerializer
)
from common.permissions import IsJefeOperaciones


class OTViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestión de Órdenes de Trabajo.
    
    Endpoints principales:
    - list: GET /ots/ - Listar OTs
    - create: POST /ots/ - Crear nueva OT
    - retrieve: GET /ots/{id}/ - Ver detalle de OT
    - update: PUT /ots/{id}/ - Actualizar OT completa
    - partial_update: PATCH /ots/{id}/ - Actualizar parcialmente
    - destroy: DELETE /ots/{id}/ - Eliminar OT (soft delete)
    
    Endpoints personalizados:
    - search: POST /ots/search/ - Búsqueda avanzada
    - search_by_container: GET /ots/search_by_container/?q=MSCU123 - Buscar por contenedor
    - search_by_bl: GET /ots/search_by_bl/?q=BL123 - Buscar por BL
    - add_container: POST /ots/{id}/add_container/ - Agregar contenedor a OT
    - remove_container: POST /ots/{id}/remove_container/ - Quitar contenedor
    - update_provision: POST /ots/{id}/update_provision/ - Actualizar provisiones
    - import_excel: POST /ots/import_excel/ - Importar OTs desde Excel
    - statistics: GET /ots/statistics/ - Estadísticas generales
    """
    
    permission_classes = [IsAuthenticated]  # Cambiado: permitir a todos los usuarios autenticados
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]  # Removed SearchFilter - we handle search in get_queryset
    filterset_fields = ['puerto_destino']  # Solo campos simples que no necesitan multi-select
    ordering_fields = ['numero_ot', 'fecha_eta', 'fecha_llegada', 'estado', 'created_at']
    ordering = ['-created_at']
    
    def get_permissions(self):
        """
        Permitir listar/retrieve a cualquier usuario autenticado.
        Otras acciones requieren IsJefeOperaciones.
        """
        if self.action in ['list', 'retrieve']:
            return [IsAuthenticated()]
        return [IsAuthenticated(), IsJefeOperaciones()]
    
    def get_queryset(self):
        """Queryset base con optimizaciones"""
        queryset = OT.objects.filter(deleted_at__isnull=True).select_related(
            'proveedor',
            'cliente',
            'modificado_por'
        )
        
        # Filtros adicionales por query params
        proveedor_id = self.request.query_params.get('proveedor_id')
        if proveedor_id:
            queryset = queryset.filter(proveedor_id=proveedor_id)
        
        # Filtro por estado - soporta múltiples valores
        estados = self.request.query_params.getlist('estado')
        if estados:
            estado_query = Q()
            for estado in estados:
                # Normalizar el estado para manejar variaciones
                estado_normalizado = estado.lower().strip()
                
                # Mapeo de variaciones comunes al valor estándar
                estado_mappings = {
                    'transito': ['transito', 'en_transito', 'tránsito', 'en tránsito', 'transit'],
                    'en_rada': ['en_rada', 'en rada', 'rada'],
                    'fact_adicionales': ['fact_adicionales', 'fact adicionales', 'facturacion adicional'],
                }
                
                # Buscar si el estado tiene variaciones conocidas
                estados_buscar = [estado_normalizado]
                for estandar, variaciones in estado_mappings.items():
                    if estado_normalizado in variaciones:
                        estados_buscar = variaciones
                        break
                
                # Agregar todas las variaciones a la consulta
                variacion_query = Q()
                for var in estados_buscar:
                    variacion_query |= Q(estado__iexact=var)
                
                estado_query |= variacion_query
            
            queryset = queryset.filter(estado_query)
        
        # Filtro por nombre de proveedor (búsqueda exacta) - soporta múltiples valores
        proveedores = self.request.query_params.getlist('proveedor')
        if proveedores:
            proveedor_query = Q()
            for proveedor in proveedores:
                proveedor_query |= Q(proveedor__nombre__iexact=proveedor)
            queryset = queryset.filter(proveedor_query)
        
        cliente_id = self.request.query_params.get('cliente_id')
        if cliente_id:
            queryset = queryset.filter(cliente_id=cliente_id)
        
        # Filtro por nombre de cliente (búsqueda exacta en original_name) - soporta múltiples valores
        clientes = self.request.query_params.getlist('cliente')
        if clientes:
            cliente_query = Q()
            for cliente in clientes:
                cliente_query |= Q(cliente__original_name__iexact=cliente)
            queryset = queryset.filter(cliente_query)
        
        # Filtro por operativo (búsqueda exacta) - soporta múltiples valores
        operativos = self.request.query_params.getlist('operativo')
        if operativos:
            operativo_query = Q()
            for operativo in operativos:
                operativo_query |= Q(operativo__iexact=operativo)
            queryset = queryset.filter(operativo_query)
        
        # Filtro por estado de provisión
        estado_provision = self.request.query_params.get('estado_provision')
        if estado_provision:
            queryset = queryset.filter(estado_provision__iexact=estado_provision)
        
        # Filtro por estado de facturación
        estado_facturado = self.request.query_params.get('estado_facturado')
        if estado_facturado:
            queryset = queryset.filter(estado_facturado__iexact=estado_facturado)
        
        # Filtro por tipo de operación
        tipo_operacion = self.request.query_params.get('tipo_operacion')
        if tipo_operacion:
            queryset = queryset.filter(tipo_operacion__iexact=tipo_operacion)
        
        # Búsqueda masiva por MBL - soporta múltiples valores
        mbls = self.request.query_params.getlist('mbl')
        if mbls:
            mbl_query = Q()
            for mbl in mbls:
                mbl_value = mbl.strip()
                if mbl_value:
                    mbl_query |= Q(master_bl__iexact=mbl_value) | Q(master_bl__icontains=mbl_value)
            if mbl_query:
                queryset = queryset.filter(mbl_query)
        
        # Búsqueda masiva por Contenedor - soporta múltiples valores
        contenedores = self.request.query_params.getlist('contenedor')
        if contenedores:
            contenedor_query = Q()
            for contenedor in contenedores:
                contenedor_value = contenedor.strip().upper()
                if contenedor_value:
                    # Buscar en el JSONField de contenedores
                    contenedor_query |= Q(contenedores__contains=[contenedor_value]) | Q(contenedores__icontains=contenedor_value)
            if contenedor_query:
                queryset = queryset.filter(contenedor_query)
        
        # Búsqueda masiva por Número de OT - soporta múltiples valores
        numeros_ot = self.request.query_params.getlist('numero_ot')
        if numeros_ot:
            ot_query = Q()
            for numero_ot in numeros_ot:
                ot_value = numero_ot.strip()
                if ot_value:
                    ot_query |= Q(numero_ot__iexact=ot_value) | Q(numero_ot__icontains=ot_value)
            if ot_query:
                queryset = queryset.filter(ot_query)
        
        # Búsqueda especial en contenedores (para el modal de asignar OT)
        search_query = self.request.query_params.get('search')
        if search_query:
            search_value = search_query.strip()
            if search_value:
                normalized_value = search_value
                normalized_upper = search_value.upper()

                # Preparar candidatos para búsqueda de contenedores (limpiando caracteres no alfanuméricos)
                container_candidates = {normalized_upper}
                if normalized_value != normalized_upper:
                    container_candidates.add(normalized_value)

                sanitized_container = re.sub(r"[^A-Z0-9]", "", normalized_upper)
                if sanitized_container and sanitized_container != normalized_upper:
                    container_candidates.add(sanitized_container)

                normalized_client_term = ' '.join(normalized_upper.split()).rstrip('.,;')

                container_filter = Q()
                for candidate in container_candidates:
                    # Búsqueda de contenedores (ahora son strings simples)
                    # Buscar tanto coincidencia exacta como parcial
                    container_filter |= Q(contenedores__contains=[candidate])
                    container_filter |= Q(contenedores__icontains=candidate)

                if not container_filter.children:
                    container_filter = Q(pk__isnull=True)

                # Búsqueda en House BLs para soportar coincidencias parciales
                house_candidates = {normalized_upper}
                if normalized_value != normalized_upper:
                    house_candidates.add(normalized_value)

                house_filter = None
                for candidate in house_candidates:
                    condition = (
                        Q(house_bls__contains=[candidate]) |
                        Q(house_bls__icontains=candidate)
                    )
                    house_filter = condition if house_filter is None else house_filter | condition

                if house_filter is None:
                    house_filter = Q(pk__isnull=True)

                client_normalized_filter = (
                    Q(cliente__normalized_name__icontains=normalized_client_term)
                    if normalized_client_term
                    else Q(pk__isnull=True)
                )

                queryset = queryset.filter(
                    Q(numero_ot__icontains=normalized_value) |
                    Q(master_bl__icontains=normalized_value) |
                    Q(notas__icontains=normalized_value) |
                    Q(cliente__original_name__icontains=normalized_value) |
                    client_normalized_filter |
                    Q(proveedor__nombre__icontains=normalized_value) |
                    Q(operativo__icontains=normalized_value) |
                    Q(barco__icontains=normalized_value) |
                    container_filter |
                    house_filter
                )
        
        return queryset
    
    def get_serializer_class(self):
        """Seleccionar serializer según acción"""
        if self.action == 'list':
            return OTListSerializer
        elif self.action == 'search':
            return OTSearchSerializer
        return OTDetailSerializer
    
    @action(detail=False, methods=['post'])
    def search(self, request):
        """
        Búsqueda avanzada de OTs.
        
        Body:
        {
            "query": "MSCU1234567",
            "search_type": "all",  // all, contenedor, master_bl, house_bl, numero_ot
            "proveedor_id": 1,     // opcional
            "cliente_id": 5,       // opcional
            "estado": "en_transito" // opcional
        }
        """
        serializer = OTSearchSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        query = serializer.validated_data['query'].upper()
        search_type = serializer.validated_data.get('search_type', 'all')
        proveedor_id = serializer.validated_data.get('proveedor_id')
        cliente_id = serializer.validated_data.get('cliente_id')
        estado = serializer.validated_data.get('estado')
        
        # Base queryset
        qs = self.get_queryset()
        
        # Aplicar filtros opcionales
        if proveedor_id:
            qs = qs.filter(proveedor_id=proveedor_id)
        if cliente_id:
            qs = qs.filter(cliente_id=cliente_id)
        if estado:
            qs = qs.filter(estado=estado)
        
        # Búsqueda según tipo
        if search_type == 'all':
            # Buscar en todos los campos
            results = qs.filter(
                Q(numero_ot__icontains=query) |
                Q(master_bl__icontains=query) |
                Q(house_bls__contains=[query]) |
                Q(contenedores__contains=[query])
            )
        elif search_type == 'contenedor':
            results = qs.filter(
                Q(contenedores__contains=[query])
            )
        elif search_type == 'master_bl':
            results = qs.filter(master_bl__icontains=query)
        elif search_type == 'house_bl':
            results = qs.filter(house_bls__contains=[query])
        elif search_type == 'numero_ot':
            results = qs.filter(numero_ot__icontains=query)
        else:
            results = qs.none()
        
        # Serializar resultados
        result_serializer = OTListSerializer(results, many=True)
        
        return Response({
            'count': results.count(),
            'search_type': search_type,
            'query': query,
            'results': result_serializer.data
        })
    
    @action(detail=False, methods=['get'])
    def search_by_container(self, request):
        """
        Buscar OT por número de contenedor.
        Query params: ?q=MSCU1234567
        """
        query = request.query_params.get('q', '').upper()
        
        if not query:
            return Response(
                {'error': 'Debe proporcionar el parámetro q con el número de contenedor'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Buscar en JSONField contenedores
        results = self.get_queryset().filter(
            Q(contenedores__contains=[query])
        )
        
        serializer = OTListSerializer(results, many=True)
        
        return Response({
            'contenedor': query,
            'count': results.count(),
            'results': serializer.data
        })
    
    @action(detail=False, methods=['get'])
    def search_by_bl(self, request):
        """
        Buscar OT por Master BL o House BL.
        Query params: ?q=BL123456&type=master (type: master o house, opcional)
        """
        query = request.query_params.get('q', '').upper()
        bl_type = request.query_params.get('type', 'both')  # master, house, both
        
        if not query:
            return Response(
                {'error': 'Debe proporcionar el parámetro q con el número de BL'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        qs = self.get_queryset()
        
        if bl_type == 'master':
            results = qs.filter(master_bl__icontains=query)
        elif bl_type == 'house':
            results = qs.filter(house_bls__contains=[query])
        else:  # both
            results = qs.filter(
                Q(master_bl__icontains=query) |
                Q(house_bls__contains=[query])
            )
        
        serializer = OTListSerializer(results, many=True)
        
        return Response({
            'bl': query,
            'type': bl_type,
            'count': results.count(),
            'results': serializer.data
        })
    
    @action(detail=True, methods=['post'])
    def add_container(self, request, pk=None):
        """Agregar un contenedor a la OT (solo se requiere el número)."""
        ot = self.get_object()

        serializer = ContenedorSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        contenedor_data = serializer.validated_data

        # Usar método del modelo
        ot.add_contenedor(contenedor_data['numero'])

        # Registrar usuario modificador
        if request.user:
            ot.modificado_por = request.user
            ot.save()

        return Response({
            'message': f'Contenedor {contenedor_data["numero"]} agregado exitosamente',
            'total_contenedores': ot.get_total_contenedores(),
            'contenedores': ot.contenedores
        })
    
    @action(detail=True, methods=['post'])
    def remove_container(self, request, pk=None):
        """
        Quitar contenedor de una OT.
        
        Body:
        {
            "numero": "MSCU1234567"
        }
        """
        ot = self.get_object()
        
        raw_numero = request.data.get('numero', '')
        numero = re.sub(r"[^A-Z0-9]", "", str(raw_numero).upper())
        if not numero:
            return Response(
                {'error': 'Debe proporcionar el número de contenedor'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Usar método del modelo
        removed = ot.remove_contenedor(numero)
        
        if removed:
            # Registrar usuario modificador
            if request.user:
                ot.modificado_por = request.user
                ot.save()
            
            return Response({
                'message': f'Contenedor {numero} removido exitosamente',
                'total_contenedores': ot.get_total_contenedores(),
                'contenedores': ot.contenedores
            })
        else:
            return Response(
                {'error': f'Contenedor {numero} no encontrado en esta OT'},
                status=status.HTTP_404_NOT_FOUND
            )
    
    @action(detail=True, methods=['post'])
    def update_provision(self, request, pk=None):
        """
        Actualizar provisiones de una OT.
        
        Body:
        {
            "items": [
                {
                    "concepto": "Flete",
                    "monto": 1500.00,
                    "categoria": "transporte"
                },
                {
                    "concepto": "Almacenaje",
                    "monto": 350.00,
                    "categoria": "puerto"
                }
            ]
        }
        """
        ot = self.get_object()
        
        serializer = ProvisionHierarchySerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        items = serializer.validated_data.get('items', [])
        
        # Usar método del modelo para recalcular total
        ot.set_provision(items)
        
        # Registrar usuario modificador
        if request.user:
            ot.modificado_por = request.user
            ot.save()
        
        return Response({
            'message': 'Provisiones actualizadas exitosamente',
            'provision_hierarchy': ot.provision_hierarchy,
            'provision_total': ot.get_provision_total()
        })
    
    @action(detail=False, methods=['post'])
    def import_excel(self, request):
        """
        Importar OTs desde múltiples archivos Excel.
        
        Body (multipart/form-data):
        - files: Array de archivos Excel (.xlsx, .xls)
        
        Respuesta:
        {
            "success": true,
            "total_rows": 15,
            "processed": 12,
            "created": 8,
            "updated": 4,
            "skipped": 3,
            "conflicts": [],  // Array de conflictos detectados (si existen)
            "errors": [],
            "message": "Importación completada exitosamente"
        }
        """
        # Obtener múltiples archivos desde request.FILES
        uploaded_files = request.FILES.getlist('files')
        
        if not uploaded_files:
            return Response(
                {'file': ['No se envió ningún archivo.']},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validar con serializer
        tipos_json = request.data.get('tipos_operacion', '[]')
        try:
            import json as json_lib
            tipos_operacion = json_lib.loads(tipos_json) if isinstance(tipos_json, str) else tipos_json
        except:
            tipos_operacion = ['importacion'] * len(uploaded_files)  # Default a importacion
        
        serializer = ExcelUploadSerializer(data={'files': uploaded_files, 'tipos_operacion': tipos_operacion})
        serializer.is_valid(raise_exception=True)
        
        uploaded_files = serializer.validated_data['files']
        tipos_operacion = serializer.validated_data.get('tipos_operacion', ['importacion'] * len(uploaded_files))
        
        # Guardar archivos temporalmente
        import tempfile
        import os
        
        temp_files = []
        try:
            # Guardar todos los archivos temporalmente
            for uploaded_file in uploaded_files:
                tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
                for chunk in uploaded_file.chunks():
                    tmp_file.write(chunk)
                tmp_file.close()
                temp_files.append((tmp_file.name, uploaded_file.name))
            
            # Procesar archivos con ExcelProcessor
            from .services.excel_processor import ExcelProcessor
            
            processor = ExcelProcessor()
            stats = processor.process_multiple_files(temp_files, tipos_operacion=tipos_operacion)
            
            # Si hay conflictos, retornarlos inmediatamente
            if stats.get('conflicts'):
                result_data = {
                    'success': False,
                    'has_conflicts': True,
                    'total_rows': stats['total_rows'],
                    'processed': 0,  # No se procesó nada aún
                    'created': 0,
                    'updated': 0,
                    'skipped': stats['skipped'],
                    'conflicts': stats['conflicts'],
                    'errors': stats['errors'],
                    'warnings': stats.get('warnings', []),
                    'message': stats.get('message', f'Se detectaron {len(stats["conflicts"])} conflictos')
                }
                
                result_serializer = ExcelImportResultSerializer(data=result_data)
                result_serializer.is_valid(raise_exception=True)
                
                return Response(
                    result_serializer.validated_data,
                    status=status.HTTP_409_CONFLICT  # 409 = Conflict
                )
            
            # No hay conflictos, procesamiento completado
            success = len(stats['errors']) == 0 or stats['processed'] > 0

            # Construir mensaje descriptivo
            message_parts = []
            if stats['processed'] > 0:
                message_parts.append(f"{stats['processed']} OTs procesadas")
            if stats['created'] > 0:
                message_parts.append(f"{stats['created']} creadas")
            if stats['updated'] > 0:
                message_parts.append(f"{stats['updated']} actualizadas")
            if stats['skipped'] > 0:
                message_parts.append(f"{stats['skipped']} omitidas")
            if stats['errors']:
                message_parts.append(f"{len(stats['errors'])} errores")

            if message_parts:
                message = 'Importación completada: ' + ', '.join(message_parts)
            else:
                message = 'No se procesaron OTs. Verifique el formato de los archivos.'

            result_data = {
                'success': success,
                'has_conflicts': False,
                'total_rows': stats['total_rows'],
                'processed': stats['processed'],
                'created': stats['created'],
                'updated': stats['updated'],
                'skipped': stats['skipped'],
                'conflicts': [],
                'errors': stats['errors'],
                'warnings': stats.get('warnings', []),
                'warnings_summary': stats.get('warnings_summary', {}),
                'message': message
            }
            
            result_serializer = ExcelImportResultSerializer(data=result_data)
            result_serializer.is_valid(raise_exception=True)
            
            return Response(
                result_serializer.validated_data,
                status=status.HTTP_200_OK if success else status.HTTP_207_MULTI_STATUS
            )
            
        except Exception as e:
            return Response(
                {
                    'success': False,
                    'has_conflicts': False,
                    'total_rows': 0,
                    'processed': 0,
                    'created': 0,
                    'updated': 0,
                    'skipped': 0,
                    'conflicts': [],
                    'errors': [{'row': 'N/A', 'error': str(e)}],
                    'message': f'Error al procesar archivos: {str(e)}'
                },
                status=status.HTTP_400_BAD_REQUEST
            )
        finally:
            # Limpiar archivos temporales
            for tmp_file_path, _ in temp_files:
                try:
                    os.unlink(tmp_file_path)
                except:
                    pass
    
    @action(detail=False, methods=['post'])
    def resolve_conflicts(self, request):
        """
        Resolver conflictos y procesar OTs con las decisiones del usuario.
        
        Body (multipart/form-data):
        - files: Array de archivos Excel
        - conflicts: JSON string con resoluciones [{"ot": "25OT221", "campo": "cliente", "resolucion": "usar_nuevo"}]
        - tipos_operacion: JSON string con tipos ["importacion", "exportacion"]
        """
        from .serializers import ConflictResolutionSerializer
        import json
        
        # Obtener archivos
        uploaded_files = request.FILES.getlist('files')
        
        if not uploaded_files:
            return Response(
                {
                    'success': False,
                    'message': 'No se proporcionaron archivos'
                },
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Obtener conflictos desde POST data (viene como JSON string en FormData)
        conflicts_json = request.POST.get('conflicts', '[]')
        try:
            conflicts_data = json.loads(conflicts_json)
        except json.JSONDecodeError:
            return Response(
                {
                    'success': False,
                    'message': 'Formato de conflictos inválido'
                },
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validar con serializer
        serializer = ConflictResolutionSerializer(data={'conflicts': conflicts_data})
        serializer.is_valid(raise_exception=True)
        
        conflicts_resolutions = serializer.validated_data['conflicts']
        
        # Obtener tipos de operación
        tipos_json = request.POST.get('tipos_operacion', '[]')
        try:
            tipos_operacion = json.loads(tipos_json) if isinstance(tipos_json, str) else tipos_json
        except:
            tipos_operacion = ['importacion'] * len(uploaded_files)
        
        if not uploaded_files:
            return Response(
                {
                    'success': False,
                    'message': 'No se proporcionaron archivos'
                },
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Guardar archivos temporalmente
        import tempfile
        import os
        
        temp_files = []
        try:
            for uploaded_file in uploaded_files:
                tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
                for chunk in uploaded_file.chunks():
                    tmp_file.write(chunk)
                tmp_file.close()
                temp_files.append((tmp_file.name, uploaded_file.name))
            
            # Re-procesar archivos con ExcelProcessor
            from .services.excel_processor import ExcelProcessor
            
            processor = ExcelProcessor()
            # Primero cargar los datos
            processor.process_multiple_files(temp_files, tipos_operacion=tipos_operacion)

            # Preparar información adicional para las resoluciones
            # Necesitamos enriquecer las resoluciones con valor_original y valor_nuevo
            for conflict_resolution in conflicts_resolutions:
                # Buscar el conflicto original en los conflictos detectados
                ot = conflict_resolution['ot']
                campo = conflict_resolution['campo']

                # Buscar en los conflictos originales para obtener los valores
                for original_conflict in processor.detected_conflicts:
                    if original_conflict['ot'] == ot and original_conflict['campo'] == campo:
                        conflict_resolution['valor_original'] = original_conflict.get('valor_actual')
                        conflict_resolution['valor_nuevo'] = original_conflict.get('valor_nuevo')
                        break

            # Luego resolver conflictos y procesar
            processed_by = request.user.username if request.user else 'system'
            stats = processor.resolve_conflicts_and_process(conflicts_resolutions, processed_by=processed_by)

            success = len(stats['errors']) == 0 or stats['processed'] > 0

            # Construir mensaje descriptivo
            message_parts = []
            if stats['processed'] > 0:
                message_parts.append(f"{stats['processed']} OTs procesadas")
            if stats['created'] > 0:
                message_parts.append(f"{stats['created']} creadas")
            if stats['updated'] > 0:
                message_parts.append(f"{stats['updated']} actualizadas")
            if stats['skipped'] > 0:
                message_parts.append(f"{stats['skipped']} omitidas")
            if stats['errors']:
                message_parts.append(f"{len(stats['errors'])} errores")

            if message_parts:
                message = 'Importación completada: ' + ', '.join(message_parts)
            else:
                message = 'No se procesaron OTs.'

            result_data = {
                'success': success,
                'has_conflicts': False,
                'total_rows': stats['total_rows'],
                'processed': stats['processed'],
                'created': stats['created'],
                'updated': stats['updated'],
                'skipped': stats['skipped'],
                'conflicts': [],
                'errors': stats['errors'],
                'warnings': stats.get('warnings', []),
                'warnings_summary': stats.get('warnings_summary', {}),
                'message': message
            }
            
            result_serializer = ExcelImportResultSerializer(data=result_data)
            result_serializer.is_valid(raise_exception=True)
            
            return Response(
                result_serializer.validated_data,
                status=status.HTTP_200_OK if success else status.HTTP_207_MULTI_STATUS
            )
            
        except Exception as e:
            return Response(
                {
                    'success': False,
                    'has_conflicts': False,
                    'total_rows': 0,
                    'processed': 0,
                    'created': 0,
                    'updated': 0,
                    'skipped': 0,
                    'conflicts': [],
                    'errors': [{'row': 'N/A', 'error': str(e)}],
                    'message': f'Error al resolver conflictos: {str(e)}'
                },
                status=status.HTTP_400_BAD_REQUEST
            )
        finally:
            for tmp_file_path, _ in temp_files:
                try:
                    os.unlink(tmp_file_path)
                except:
                    pass
    
    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """
        Estadísticas generales de OTs.
        
        Retorna:
        - Total de OTs por estado
        - Total de contenedores
        - Total de provisiones
        - OTs por proveedor
        """
        qs = self.get_queryset()
        
        # Por estado
        by_estado = {}
        for estado_code, estado_label in OT.STATUS_CHOICES:
            count = qs.filter(estado=estado_code).count()
            by_estado[estado_code] = {
                'label': estado_label,
                'count': count
            }
        
        from django.db.models.functions import Cast, Coalesce, JsonbArrayLength
        from django.contrib.postgres.fields.jsonb import KeyTextTransform

        # Total de OTs
        total_ots = qs.count()

        # Total de contenedores usando agregación de base de datos
        total_contenedores = qs.aggregate(
            total=Coalesce(Sum(JsonbArrayLength('contenedores')), 0)
        )['total']

        # Total de provisiones usando agregación de base de datos
        total_provision = qs.aggregate(
            total=Coalesce(Sum(Cast(KeyTextTransform('total', 'provision_hierarchy'), models.FloatField())), 0.0)
        )['total']
        
        # Por proveedor (top 10)
        by_proveedor = qs.values(
            'proveedor__nombre'
        ).annotate(
            count=Count('id')
        ).order_by('-count')[:10]
        
        return Response({
            'total_ots': total_ots,
            'total_contenedores': total_contenedores,
            'total_provision': float(total_provision),
            'by_estado': by_estado,
            'top_proveedores': list(by_proveedor)
        })
    
    @action(detail=False, methods=['get'], url_path='cards-stats')
    def cards_stats(self, request):
        """
        Estadísticas específicas para las cards del dashboard.
        Respeta los filtros aplicados y calcula:
        - Total de OTs filtradas
        - OTs Facturadas (con fecha_recepcion_factura)
        - OTs Cerradas (estado='cerrada')
        - OTs Pendientes de Cierre (estado='finalizada')
        
        Usa agregaciones de base de datos para ser eficiente incluso con muchos registros.
        """
        # Usar el mismo queryset filtrado que usa get_queryset()
        qs = self.get_queryset()
        
        # Total de OTs con los filtros aplicados
        total = qs.count()
        
        # OTs Facturadas: las que tienen fecha_recepcion_factura (DateField, solo puede ser NULL o fecha válida)
        facturadas = qs.exclude(fecha_recepcion_factura__isnull=True).count()
        
        # OTs Cerradas: estado = 'cerrada'
        cerradas = qs.filter(estado__iexact='cerrada').count()
        
        # OTs Pendientes de Cierre: estado = 'finalizada'
        pendientes_cierre = qs.filter(estado__iexact='finalizada').count()
        
        return Response({
            'total': total,
            'facturadas': facturadas,
            'cerradas': cerradas,
            'pendientes_cierre': pendientes_cierre,
        })
    
    @action(detail=False, methods=['post'], url_path='import-provision-acajutla')
    def import_provision_acajutla(self, request):
        """
        Importar CSV de Provisión Acajutla.
        Solo actualiza campos: fecha_provision y barco
        Respeta prioridades: MANUAL > CSV > EXCEL
        
        Formato esperado del CSV:
        - Columna "OT" (índice 6)
        - Columna "BARCO" (índice 13)
        - Columna "FECHA DE PROVISION" (índice 16)
        
        Valores ignorados en FECHA DE PROVISION:
        - "N/A"
        - "SOLICITUD DE PAGO"
        - Cualquier texto que no sea fecha
        """
        import csv
        import io
        from datetime import datetime
        
        if 'file' not in request.FILES:
            return Response(
                {'error': 'No se proporcionó ningún archivo'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        csv_file = request.FILES['file']
        
        # Validar que sea CSV
        if not csv_file.name.endswith('.csv'):
            return Response(
                {'error': 'El archivo debe ser CSV'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Leer CSV
            decoded_file = csv_file.read().decode('utf-8-sig')  # utf-8-sig para manejar BOM
            csv_reader = csv.reader(io.StringIO(decoded_file))
            
            stats = {
                'processed': 0,
                'updated': 0,
                'skipped': 0,
                'errors': []
            }
            
            # Saltar header (primeras 2 líneas)
            next(csv_reader, None)
            next(csv_reader, None)
            
            for row_num, row in enumerate(csv_reader, start=3):
                if len(row) < 17:
                    continue  # Fila incompleta
                
                try:
                    ot_numero = row[6].strip() if len(row) > 6 else None
                    barco_csv = row[13].strip() if len(row) > 13 else None
                    fecha_provision_str = row[16].strip() if len(row) > 16 else None
                    
                    # Validar que tengamos al menos el número de OT
                    if not ot_numero or ot_numero == 'PLG SV':
                        continue
                    
                    # Buscar la OT
                    try:
                        ot = OT.objects.get(numero_ot__iexact=ot_numero, deleted_at__isnull=True)
                    except OT.DoesNotExist:
                        stats['skipped'] += 1
                        continue
                    except OT.MultipleObjectsReturned:
                        stats['errors'].append({
                            'row': row_num,
                            'ot': ot_numero,
                            'error': 'Múltiples OTs con el mismo número'
                        })
                        continue
                    
                    updated = False
                    
                    # Procesar BARCO
                    if barco_csv and barco_csv != '-' and barco_csv != 'N/A':
                        if ot.can_update_field('barco', 'csv'):
                            ot.barco = barco_csv
                            ot.barco_source = 'csv'
                            updated = True
                    
                    # Procesar FECHA DE PROVISION
                    if fecha_provision_str and fecha_provision_str not in ['N/A', 'SOLICITUD DE PAGO', '-', '']:
                        # Intentar parsear la fecha
                        try:
                            # Formatos comunes: D/M/YYYY, DD/MM/YYYY
                            for date_format in ['%d/%m/%Y', '%m/%d/%Y', '%Y-%m-%d']:
                                try:
                                    fecha_obj = datetime.strptime(fecha_provision_str, date_format).date()
                                    break
                                except ValueError:
                                    continue
                            else:
                                # No se pudo parsear
                                raise ValueError(f"Formato de fecha no reconocido: {fecha_provision_str}")
                            
                            if ot.can_update_field('fecha_provision', 'csv'):
                                ot.fecha_provision = fecha_obj
                                ot.provision_source = 'csv'
                                updated = True
                        
                        except ValueError as e:
                            # No es una fecha válida, ignorar
                            pass
                    
                    if updated:
                        ot.save()
                        stats['updated'] += 1
                    
                    stats['processed'] += 1
                
                except Exception as e:
                    stats['errors'].append({
                        'row': row_num,
                        'ot': ot_numero if 'ot_numero' in locals() else 'N/A',
                        'error': str(e)
                    })
            
            return Response({
                'message': f'Provisión Acajutla importada: {stats["updated"]} OTs actualizadas de {stats["processed"]} procesadas',
                'stats': stats
            }, status=status.HTTP_200_OK)
        
        except Exception as e:
            return Response(
                {'error': f'Error procesando CSV: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'], url_path='filter-values')
    def filter_values(self, request):
        """
        Returns unique values for the filterable fields.
        """
        qs = self.get_queryset()
        
        clientes = qs.values_list('cliente__original_name', flat=True).distinct().order_by('cliente__original_name')
        operativos = qs.values_list('operativo', flat=True).distinct().order_by('operativo')
        proveedores = qs.values_list('proveedor__nombre', flat=True).distinct().order_by('proveedor__nombre')
        estados = qs.values_list('estado', flat=True).distinct().order_by('estado')
        estados_provision = qs.values_list('estado_provision', flat=True).distinct().order_by('estado_provision')
        estados_facturado = qs.values_list('estado_facturado', flat=True).distinct().order_by('estado_facturado')

        return Response({
            'clientes': [c for c in clientes if c],
            'operativos': [o for o in operativos if o],
            'proveedores': [p for p in proveedores if p],
            'estados': [e for e in estados if e],
            'estados_provision': [ep for ep in estados_provision if ep],
            'estados_facturado': [ef for ef in estados_facturado if ef],
        })

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        """
        Exportar OTs a Excel con formato profesional.
        Respeta todos los filtros aplicados en get_queryset().
        EXPORTA TODOS LOS REGISTROS FILTRADOS (no solo la página actual).

        Query params: Los mismos que el listado (estado, proveedor, cliente, etc.)

        Retorna un archivo Excel con:
        - Fechas en formato dd/mm/yyyy
        - Montos con formato contable
        - Encabezados con estilo profesional
        - Anchos de columna ajustados
        - Colores y formato de tabla
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse
        from datetime import datetime

        # MEMORY OPTIMIZATION: Use iterator instead of loading all records
        # Previous: Load 10,000 records = ~500MB RAM
        # New: Process in chunks = ~50MB RAM
        queryset = self.filter_queryset(self.get_queryset()).select_related(
            'proveedor', 'cliente', 'modificado_por'
        )

        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "OTs"

        # Definir headers
        headers = [
            'Número OT', 'Estado', 'Cliente', 'Operativo', 'MBL', 'Contenedores',
            'Naviera', 'Barco', 'Fecha Provisión', 'Fecha Facturación',
            'Tipo Embarque', 'Puerto Origen', 'Puerto Destino', 'ETD', 'ETA',
            'ETA Confirmada', 'House BLs', 'Estado Provisión', 'Estado Facturado',
            'Express Release', 'Contra Entrega', 'Solicitud Facturación',
            'Envío Cierre OT', 'Fecha Creación', 'Última Actualización', 'Comentarios'
        ]

        # Estilos profesionales
        header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")  # Azul oscuro
        header_font = Font(color="FFFFFF", bold=True, size=12, name='Calibri')
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Estilo para filas alternas
        alt_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")  # Gris claro
        
        # Bordes
        thin_border = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='thin', color='D1D5DB')
        )
        
        # Fuente para datos
        data_font = Font(size=11, name='Calibri')
        data_alignment = Alignment(horizontal="left", vertical="center")

        # Escribir encabezados
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        # Escribir datos usando iterator (memory efficient)
        row_num = 2
        for ot in queryset.iterator(chunk_size=100):
            # Formatear contenedores
            contenedores_str = ', '.join(ot.get_contenedores_numeros()) if ot.contenedores else ''

            # Formatear House BLs
            house_bls_str = ', '.join(ot.house_bls) if ot.house_bls else ''

            row_data = [
                ot.numero_ot or '',
                ot.get_estado_display() or '',
                ot.cliente.original_name if ot.cliente else '',
                ot.operativo or '',
                ot.master_bl or '',
                contenedores_str,
                ot.proveedor.nombre if ot.proveedor else '',
                ot.barco or '',
                ot.fecha_provision,
                ot.fecha_recepcion_factura,
                ot.tipo_embarque or '',
                ot.puerto_origen or '',
                ot.puerto_destino or '',
                ot.etd,
                ot.fecha_eta,
                ot.fecha_llegada,
                house_bls_str,
                ot.get_estado_provision_display() or '',
                ot.get_estado_facturado_display() or '',
                ot.express_release_fecha,
                ot.contra_entrega_fecha,
                ot.fecha_solicitud_facturacion,
                ot.envio_cierre_ot,
                ot.created_at.date() if ot.created_at else None,
                ot.updated_at.date() if ot.updated_at else None,
                ot.comentarios or ''
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = thin_border
                cell.font = data_font
                cell.alignment = data_alignment
                
                # Aplicar fondo alterno para mejor legibilidad
                if row_num % 2 == 0:
                    cell.fill = alt_fill

                # Aplicar formato de fecha para columnas de fecha (DD/MM/YYYY)
                if col_num in [9, 10, 14, 15, 16, 20, 21, 22, 23, 24, 25]:  # Columnas de fecha
                    if value:
                        cell.number_format = 'DD/MM/YYYY'
                        cell.alignment = Alignment(horizontal="center", vertical="center")

            row_num += 1

        # Ajustar anchos de columna
        column_widths = {
            1: 15,   # Número OT
            2: 18,   # Estado
            3: 30,   # Cliente
            4: 20,   # Operativo
            5: 20,   # MBL
            6: 40,   # Contenedores
            7: 25,   # Naviera
            8: 25,   # Barco
            9: 15,   # Fecha Provisión
            10: 15,  # Fecha Facturación
            11: 15,  # Tipo Embarque
            12: 25,  # Puerto Origen
            13: 25,  # Puerto Destino
            14: 12,  # ETD
            15: 12,  # ETA
            16: 15,  # ETA Confirmada
            17: 35,  # House BLs
            18: 18,  # Estado Provisión
            19: 18,  # Estado Facturado
            20: 15,  # Express Release
            21: 15,  # Contra Entrega
            22: 18,  # Solicitud Facturación
            23: 15,  # Envío Cierre OT
            24: 15,  # Fecha Creación
            25: 18,  # Última Actualización
            26: 40   # Comentarios
        }

        for col_num, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col_num)].width = width

        # Congelar primera fila
        ws.freeze_panes = 'A2'

        # Agregar filtros de Excel (autofilter)
        ws.auto_filter.ref = ws.dimensions
        
        # Preparar respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        total_records = queryset.count()
        filename = f'OTs_{total_records}_registros_{timestamp}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response['Access-Control-Expose-Headers'] = 'Content-Disposition'

        # Guardar workbook en la respuesta
        wb.save(response)

        return response
